import datetime
import io
import json

import openpyxl
from dateutil import parser

from app.export_service import ExportService
from tests.base_test_questionnaire import BaseTestQuestionnaire


class TestQuestionnaire(BaseTestQuestionnaire):
    def test_base_endpoint(self):
        rv = self.client.get("/", follow_redirects=True, content_type="application/json")
        self.assert_success(rv)
        response = rv.json

        endpoints = [
            ("api.categorybyresourceendpoint", "/api/resource/<resource_id>/category"),
            ("api.categorybystudyendpoint", "/api/study/<study_id>/category"),
            ("api.categoryendpoint", "/api/category/<id>"),
            ("api.categorylistendpoint", "/api/category"),
            ("api.questionnaireendpoint", "/api/q/<name>/<id>"),
            ("api.resourcebycategoryendpoint", "/api/category/<category_id>/resource"),
            ("api.resourcecategoryendpoint", "/api/resource_category/<id>"),
            ("api.resourcecategorylistendpoint", "/api/resource_category"),
            ("api.resourceendpoint", "/api/resource/<id>"),
            ("api.resourcelistendpoint", "/api/resource"),
            ("api.rootcategorylistendpoint", "/api/category/root"),
            ("api.sessionendpoint", "/api/session"),
            ("api.studybycategoryendpoint", "/api/category/<category_id>/study"),
            ("api.studycategoryendpoint", "/api/study_category/<id>"),
            ("api.studycategorylistendpoint", "/api/study_category"),
            ("api.studyendpoint", "/api/study/<id>"),
            ("api.studylistendpoint", "/api/study"),
            ("api.userendpoint", "/api/user/<id>"),
            ("api.userlistendpoint", "/api/user"),
            ("api.zipcodecoordsendpoint", "/api/zip_code_coords/<id>"),
            ("auth.forgot_password", "/api/forgot_password"),
            ("auth.login_password", "/api/login_password"),
            ("auth.reset_password", "/api/reset_password"),
        ]

        for endpoint in endpoints:
            self.assertEqual(response[endpoint[0]], endpoint[1])

    def test_questionnare_post_fails_if_flow_does_not_exist(self):
        evaluation_history_self_questionnaire = {"self_identifies_autistic": True, "years_old_at_first_diagnosis": 5}
        rv = self.client.post(
            "api/flow/noSuchFlow/evaluation_history_self_questionnaire",
            data=self.jsonify(evaluation_history_self_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assertEqual(
            404,
            rv.status_code,
            "This endpoint should require that the flow exists and that the question is in the flow.",
        )
        response = rv.json
        self.assertIsNotNone(response)
        self.assertEqual(
            "Unknown path.", response["message"], "There should be a clear error message explaining what went wrong."
        )

    def test_questionnare_post_fails_if_question_not_in_flow(self):
        evaluation_history_self_questionnaire = {"self_identifies_autistic": True, "years_old_at_first_diagnosis": 5}
        rv = self.client.post(
            "api/flow/self_intake/guardian_demographics_questionnaire",
            data=self.jsonify(evaluation_history_self_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assertEqual(
            400,
            rv.status_code,
            "This endpoint should require that the flow exists and that the question is in the flow.",
        )
        response = rv.json
        self.assertIsNotNone(response)
        self.assertEqual(
            "not_in_the_flow", response["code"], "There should be a clear error message explaining what went wrong."
        )

    def test_questionnaire_post_fails_if_user_not_connected_to_participant(self):
        cq = {"first_name": "Darah", "marketing_channel": "Subway sign"}
        rv = self.client.post(
            "api/flow/self_intake/contact_questionnaire",
            data=self.jsonify(cq),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assertEqual(
            400, rv.status_code, "This endpoint should require a participant id that is associated with current user."
        )
        response = rv.json
        self.assertIsNotNone(response)
        self.assertEqual(
            "Unable to save the provided object.",
            response["message"],
            "There should be a clear error message explaining what went wrong.",
        )

    def test_questionnionare_post_creates_log_record(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        cq = {"first_name": "Darah", "marketing_channel": "Subway sign", "participant_id": p.id}
        rv = self.client.post(
            "api/flow/self_intake/contact_questionnaire",
            data=self.jsonify(cq),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        from app.models import StepLog

        log = self.session.query(StepLog).all()
        self.assertIsNotNone(log)
        self.assertTrue(len(log) > 0)

    def test_clinical_diagnoses_questionnaire_basics(self):
        self.construct_clinical_diagnoses_questionnaire()
        from app.models import ClinicalDiagnosesQuestionnaire

        cq = self.session.query(ClinicalDiagnosesQuestionnaire).first()
        self.assertIsNotNone(cq)
        cq_id = cq.id
        rv = self.client.get(
            "/api/q/clinical_diagnoses_questionnaire/%i" % cq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], cq_id)
        self.assertEqual(response["medical"], cq.medical)
        self.assertEqual(response["genetic"], cq.genetic)

    def test_modify_clinical_diagnoses_questionnaire_basics(self):
        self.construct_clinical_diagnoses_questionnaire()
        from app.models import ClinicalDiagnosesQuestionnaire

        cq = self.session.query(ClinicalDiagnosesQuestionnaire).first()
        self.assertIsNotNone(cq)
        cq_id = cq.id
        rv = self.client.get(
            "/api/q/clinical_diagnoses_questionnaire/%i" % cq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        response["developmental"] = ["intellectual"]
        response["mental_health"] = ["depression"]
        response["medical"] = ["gastrointestinal"]
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/clinical_diagnoses_questionnaire/%i" % cq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/clinical_diagnoses_questionnaire/%i" % cq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["developmental"], ["intellectual"])
        self.assertEqual(response["mental_health"], ["depression"])
        self.assertEqual(response["medical"], ["gastrointestinal"])
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_clinical_diagnoses_questionnaire(self):
        cq = self.construct_clinical_diagnoses_questionnaire()
        cq_id = cq.id
        rv = self.client.get(
            "api/q/clinical_diagnoses_questionnaire/%i" % cq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/clinical_diagnoses_questionnaire/%i" % cq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/clinical_diagnoses_questionnaire/%i" % cq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_clinical_diagnoses_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        clinical_diagnoses_questionnaire = {"medical": ["seizure"], "genetic": ["fragileX"], "participant_id": p.id}
        rv = self.client.post(
            "api/flow/self_intake/clinical_diagnoses_questionnaire",
            data=self.jsonify(clinical_diagnoses_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["medical"], ["seizure"])
        self.assertEqual(response["genetic"], ["fragileX"])
        self.assertIsNotNone(response["id"])

    def test_contact_questionnaire_basics(self):
        self.construct_contact_questionnaire()
        from app.models import ContactQuestionnaire

        cq = self.session.query(ContactQuestionnaire).first()
        self.assertIsNotNone(cq)
        cq_id = cq.id
        rv = self.client.get(
            "/api/q/contact_questionnaire/%i" % cq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], cq_id)
        self.assertEqual(response["phone"], cq.phone)
        self.assertEqual(response["marketing_channel"], cq.marketing_channel)

    def test_modify_contact_questionnaire_basics(self):
        self.construct_contact_questionnaire()
        from app.models import ContactQuestionnaire

        cq = self.session.query(ContactQuestionnaire).first()
        self.assertIsNotNone(cq)
        cq_id = cq.id
        rv = self.client.get(
            "/api/q/contact_questionnaire/%i" % cq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        response = rv.json
        response["phone"] = "123-456-7890"
        response["zip"] = 22345
        response["marketing_channel"] = "flyer"
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/contact_questionnaire/%i" % cq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/contact_questionnaire/%i" % cq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["phone"], "123-456-7890")
        self.assertEqual(response["zip"], 22345)
        self.assertEqual(response["marketing_channel"], "flyer")
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_contact_questionnaire(self):
        cq = self.construct_contact_questionnaire()
        cq_id = cq.id
        rv = self.client.get(
            "api/q/contact_questionnaire/%i" % cq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/contact_questionnaire/%i" % cq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/contact_questionnaire/%i" % cq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assertEqual(404, rv.status_code)

    def test_create_contact_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        contact_questionnaire = {"phone": "123-456-7890", "marketing_channel": "Subway sign", "participant_id": p.id}
        rv = self.client.post(
            "api/flow/self_intake/contact_questionnaire",
            data=self.jsonify(contact_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["phone"], "123-456-7890")
        self.assertEqual(response["marketing_channel"], "Subway sign")
        self.assertIsNotNone(response["id"])

    def test_current_behaviors_dependent_questionnaire_basics(self):
        self.construct_current_behaviors_dependent_questionnaire()
        from app.models import CurrentBehaviorsDependentQuestionnaire

        cbdq = self.session.query(CurrentBehaviorsDependentQuestionnaire).first()
        self.assertIsNotNone(cbdq)
        cbdq_id = cbdq.id
        rv = self.client.get(
            "/api/q/current_behaviors_dependent_questionnaire/%i" % cbdq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], cbdq_id)
        self.assertEqual(response["concerning_behaviors"], cbdq.concerning_behaviors)
        self.assertEqual(response["has_academic_difficulties"], cbdq.has_academic_difficulties)

    def test_modify_current_behaviors_dependent_questionnaire_basics(self):
        self.construct_current_behaviors_dependent_questionnaire()
        from app.models import CurrentBehaviorsDependentQuestionnaire

        cbdq = self.session.query(CurrentBehaviorsDependentQuestionnaire).first()
        self.assertIsNotNone(cbdq)
        cbdq_id = cbdq.id
        rv = self.client.get(
            "/api/q/current_behaviors_dependent_questionnaire/%i" % cbdq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        response["dependent_verbal_ability"] = "nonVerbal"
        response["concerning_behaviors"] = ["elopement"]
        response["has_academic_difficulties"] = False
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/current_behaviors_dependent_questionnaire/%i" % cbdq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/current_behaviors_dependent_questionnaire/%i" % cbdq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["dependent_verbal_ability"], "nonVerbal")
        self.assertEqual(response["concerning_behaviors"], ["elopement"])
        self.assertEqual(response["has_academic_difficulties"], False)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_current_behaviors_dependent_questionnaire(self):
        cbdq = self.construct_current_behaviors_dependent_questionnaire()
        cbdq_id = cbdq.id
        rv = self.client.get(
            "api/q/current_behaviors_dependent_questionnaire/%i" % cbdq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/current_behaviors_dependent_questionnaire/%i" % cbdq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/current_behaviors_dependent_questionnaire/%i" % cbdq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_current_behaviors_dependent_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.dependent)
        headers = self.logged_in_headers(u)

        current_behaviors_dependent_questionnaire = {
            "dependent_verbal_ability": "verbal, AACsystem",
            "has_academic_difficulties": False,
            "participant_id": p.id,
        }
        rv = self.client.post(
            "api/flow/dependent_intake/current_behaviors_dependent_questionnaire",
            data=self.jsonify(current_behaviors_dependent_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["dependent_verbal_ability"], "verbal, AACsystem")
        self.assertEqual(response["has_academic_difficulties"], False)
        self.assertIsNotNone(response["id"])

    def test_current_behaviors_self_questionnaire_basics(self):
        self.construct_current_behaviors_self_questionnaire()
        from app.models import CurrentBehaviorsSelfQuestionnaire

        cbsq = self.session.query(CurrentBehaviorsSelfQuestionnaire).first()
        self.assertIsNotNone(cbsq)
        cbsq_id = cbsq.id
        rv = self.client.get(
            "/api/q/current_behaviors_self_questionnaire/%i" % cbsq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], cbsq_id)
        self.assertEqual(response["has_academic_difficulties"], cbsq.has_academic_difficulties)

    def test_modify_current_behaviors_self_questionnaire_basics(self):
        self.construct_current_behaviors_self_questionnaire()
        from app.models import CurrentBehaviorsSelfQuestionnaire

        cbsq = self.session.query(CurrentBehaviorsSelfQuestionnaire).first()
        self.assertIsNotNone(cbsq)
        cbsq_id = cbsq.id
        rv = self.client.get(
            "/api/q/current_behaviors_self_questionnaire/%i" % cbsq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        response["self_verbal_ability"] = ["nonVerbal"]
        response["academic_difficulty_areas"] = ["math"]
        response["has_academic_difficulties"] = False
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/current_behaviors_self_questionnaire/%i" % cbsq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/current_behaviors_self_questionnaire/%i" % cbsq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["self_verbal_ability"], ["nonVerbal"])
        self.assertEqual(response["academic_difficulty_areas"], ["math"])
        self.assertEqual(response["has_academic_difficulties"], False)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_current_behaviors_self_questionnaire(self):
        cbsq = self.construct_current_behaviors_self_questionnaire()
        cbsq_id = cbsq.id
        rv = self.client.get(
            "api/q/current_behaviors_self_questionnaire/%i" % cbsq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/current_behaviors_self_questionnaire/%i" % cbsq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/current_behaviors_self_questionnaire/%i" % cbsq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_current_behaviors_self_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        current_behaviors_self_questionnaire = {
            "self_verbal_ability": ["verbal", "AACsystem"],
            "has_academic_difficulties": False,
            "participant_id": p.id,
        }
        rv = self.client.post(
            "api/flow/self_intake/current_behaviors_self_questionnaire",
            data=self.jsonify(current_behaviors_self_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["self_verbal_ability"], ["verbal", "AACsystem"])
        self.assertEqual(response["has_academic_difficulties"], False)
        self.assertIsNotNone(response["id"])

    def test_demographics_questionnaire_basics(self):
        self.construct_demographics_questionnaire()
        from app.models import DemographicsQuestionnaire

        dq = self.session.query(DemographicsQuestionnaire).first()
        self.assertIsNotNone(dq)
        dq_id = dq.id
        rv = self.client.get(
            "/api/q/demographics_questionnaire/%i" % dq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], dq_id)
        self.assertEqual(response["birth_sex"], dq.birth_sex)
        self.assertEqual(response["gender_identity"], dq.gender_identity)

    def test_modify_demographics_questionnaire_basics(self):
        self.construct_demographics_questionnaire()
        from app.models import DemographicsQuestionnaire

        dq = self.session.query(DemographicsQuestionnaire).first()
        self.assertIsNotNone(dq)
        dq_id = dq.id
        rv = self.client.get(
            "/api/q/demographics_questionnaire/%i" % dq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        response["gender_identity"] = "genderOther"
        response["race_ethnicity"] = ["raceOther"]
        u2 = self.construct_user(email="rainbows@rainy.com")
        response["user_id"] = u2.id
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/demographics_questionnaire/%i" % dq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/demographics_questionnaire/%i" % dq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["gender_identity"], "genderOther")
        self.assertEqual(response["race_ethnicity"], ["raceOther"])
        self.assertEqual(response["user_id"], u2.id)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_demographics_questionnaire(self):
        dq = self.construct_demographics_questionnaire()
        dq_id = dq.id
        rv = self.client.get(
            "api/q/demographics_questionnaire/%i" % dq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/demographics_questionnaire/%i" % dq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/demographics_questionnaire/%i" % dq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_demographics_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        demographics_questionnaire = {"birth_sex": "female", "gender_identity": "genderOther", "participant_id": p.id}
        rv = self.client.post(
            "api/flow/self_intake/demographics_questionnaire",
            data=self.jsonify(demographics_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["birth_sex"], "female")
        self.assertEqual(response["gender_identity"], "genderOther")
        self.assertIsNotNone(response["id"])

    def test_developmental_questionnaire_basics(self):
        self.construct_developmental_questionnaire()
        from app.models import DevelopmentalQuestionnaire

        dq = self.session.query(DevelopmentalQuestionnaire).first()
        self.assertIsNotNone(dq)
        dq_id = dq.id
        rv = self.client.get(
            "/api/q/developmental_questionnaire/%i" % dq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], dq_id)
        self.assertEqual(response["had_birth_complications"], dq.had_birth_complications)
        self.assertEqual(response["when_motor_milestones"], dq.when_motor_milestones)

    def test_modify_developmental_questionnaire_basics(self):
        self.construct_developmental_questionnaire()
        from app.models import DevelopmentalQuestionnaire

        dq = self.session.query(DevelopmentalQuestionnaire).first()
        self.assertIsNotNone(dq)
        dq_id = dq.id
        rv = self.client.get(
            "/api/q/developmental_questionnaire/%i" % dq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        response["when_motor_milestones"] = "notYet"
        response["when_language_milestones"] = "notYet"
        response["when_toileting_milestones"] = "early"
        u2 = self.construct_user(email="rainbows@rainy.com")
        response["user_id"] = u2.id
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/developmental_questionnaire/%i" % dq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/developmental_questionnaire/%i" % dq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["when_motor_milestones"], "notYet")
        self.assertEqual(response["when_language_milestones"], "notYet")
        self.assertEqual(response["when_toileting_milestones"], "early")
        self.assertEqual(response["user_id"], u2.id)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_developmental_questionnaire(self):
        dq = self.construct_developmental_questionnaire()
        dq_id = dq.id
        rv = self.client.get(
            "api/q/developmental_questionnaire/%i" % dq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/developmental_questionnaire/%i" % dq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/developmental_questionnaire/%i" % dq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_developmental_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.dependent)
        headers = self.logged_in_headers(u)

        developmental_questionnaire = {
            "had_birth_complications": True,
            "birth_complications_description": "C-Section",
            "participant_id": p.id,
        }
        rv = self.client.post(
            "api/flow/dependent_intake/developmental_questionnaire",
            data=self.jsonify(developmental_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["had_birth_complications"], True)
        self.assertEqual(response["birth_complications_description"], "C-Section")
        self.assertIsNotNone(response["id"])

    def test_education_dependent_questionnaire_basics(self):
        self.construct_education_dependent_questionnaire()
        from app.models import EducationDependentQuestionnaire

        eq = self.session.query(EducationDependentQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.client.get(
            "/api/q/education_dependent_questionnaire/%i" % eq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], eq_id)
        self.assertEqual(response["school_name"], eq.school_name)
        self.assertEqual(response["school_type"], eq.school_type)

    def test_modify_education_dependent_questionnaire_basics(self):
        self.construct_education_dependent_questionnaire()
        from app.models import EducationDependentQuestionnaire

        eq = self.session.query(EducationDependentQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.client.get(
            "/api/q/education_dependent_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        response["school_name"] = "Sesame School"
        response["school_type"] = "public"
        response["dependent_placement"] = "vocational"
        u2 = self.construct_user(email="rainbows@rainy.com")
        response["user_id"] = u2.id
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/education_dependent_questionnaire/%i" % eq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/education_dependent_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["school_name"], "Sesame School")
        self.assertEqual(response["school_type"], "public")
        self.assertEqual(response["dependent_placement"], "vocational")
        self.assertEqual(response["user_id"], u2.id)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_education_dependent_questionnaire(self):
        eq = self.construct_education_dependent_questionnaire()
        eq_id = eq.id
        rv = self.client.get(
            "api/q/education_dependent_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/education_dependent_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/education_dependent_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_education_dependent_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        education_dependent_questionnaire = {
            "attends_school": True,
            "school_name": "Attreyu Academy",
            "participant_id": p.id,
        }
        rv = self.client.post(
            "api/flow/dependent_intake/education_dependent_questionnaire",
            data=self.jsonify(education_dependent_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["attends_school"], True)
        self.assertEqual(response["school_name"], "Attreyu Academy")
        self.assertIsNotNone(response["id"])

    def test_education_self_questionnaire_basics(self):
        self.construct_education_self_questionnaire()
        from app.models import EducationSelfQuestionnaire

        eq = self.session.query(EducationSelfQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.client.get(
            "/api/q/education_self_questionnaire/%i" % eq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], eq_id)
        self.assertEqual(response["school_name"], eq.school_name)
        self.assertEqual(response["school_type"], eq.school_type)

    def test_modify_education_self_questionnaire_basics(self):
        self.construct_education_self_questionnaire()
        from app.models import EducationSelfQuestionnaire

        eq = self.session.query(EducationSelfQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.client.get(
            "/api/q/education_self_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        response["school_name"] = "Sesame School"
        response["school_type"] = "public"
        response["self_placement"] = "vocational"
        u2 = self.construct_user(email="rainbows@rainy.com")
        response["user_id"] = u2.id
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/education_self_questionnaire/%i" % eq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/education_self_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["school_name"], "Sesame School")
        self.assertEqual(response["school_type"], "public")
        self.assertEqual(response["self_placement"], "vocational")
        self.assertEqual(response["user_id"], u2.id)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_education_self_questionnaire(self):
        eq = self.construct_education_self_questionnaire()
        eq_id = eq.id
        rv = self.client.get(
            "api/q/education_self_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/education_self_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/education_self_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_education_self_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        education_self_questionnaire = {
            "attends_school": True,
            "school_name": "Attreyu Academy",
            "participant_id": p.id,
        }
        rv = self.client.post(
            "api/flow/self_intake/education_self_questionnaire",
            data=self.jsonify(education_self_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["attends_school"], True)
        self.assertEqual(response["school_name"], "Attreyu Academy")
        self.assertIsNotNone(response["id"])

    def test_employment_questionnaire_basics(self):
        self.construct_employment_questionnaire()
        from app.models import EmploymentQuestionnaire

        eq = self.session.query(EmploymentQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.client.get(
            "/api/q/employment_questionnaire/%i" % eq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], eq_id)
        self.assertEqual(response["is_currently_employed"], eq.is_currently_employed)
        self.assertEqual(response["employment_capacity"], eq.employment_capacity)
        self.assertEqual(response["has_employment_support"], eq.has_employment_support)

    def test_modify_employment_questionnaire_basics(self):
        self.construct_employment_questionnaire()
        from app.models import EmploymentQuestionnaire

        eq = self.session.query(EmploymentQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.client.get(
            "/api/q/employment_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        response["is_currently_employed"] = False
        response["employment_capacity"] = None
        response["has_employment_support"] = "yes"
        u2 = self.construct_user(email="rainbows@rainy.com")
        response["user_id"] = u2.id
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/employment_questionnaire/%i" % eq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/employment_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["is_currently_employed"], False)
        self.assertEqual(response["employment_capacity"], None)
        self.assertEqual(response["has_employment_support"], "yes")
        self.assertEqual(response["user_id"], u2.id)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_employment_questionnaire(self):
        eq = self.construct_employment_questionnaire()
        eq_id = eq.id
        rv = self.client.get(
            "api/q/employment_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/employment_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/employment_questionnaire/%i" % eq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_employment_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        employment_questionnaire = {
            "is_currently_employed": True,
            "employment_capacity": "partTime",
            "participant_id": p.id,
        }
        rv = self.client.post(
            "api/flow/self_intake/employment_questionnaire",
            data=self.jsonify(employment_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["is_currently_employed"], True)
        self.assertEqual(response["employment_capacity"], "partTime")
        self.assertIsNotNone(response["id"])

    def test_evaluation_history_dependent_questionnaire_basics(self):
        self.construct_evaluation_history_dependent_questionnaire()
        from app.models import (
            EvaluationHistoryDependentQuestionnaire,
        )

        ehq = self.session.query(EvaluationHistoryDependentQuestionnaire).first()
        self.assertIsNotNone(ehq)
        ehq_id = ehq.id
        rv = self.client.get(
            "/api/q/evaluation_history_dependent_questionnaire/%i" % ehq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], ehq_id)
        self.assertEqual(response["self_identifies_autistic"], ehq.self_identifies_autistic)
        self.assertEqual(response["years_old_at_first_diagnosis"], ehq.years_old_at_first_diagnosis)

    def test_modify_evaluation_history_dependent_questionnaire_basics(self):
        self.construct_evaluation_history_dependent_questionnaire()
        from app.models import (
            EvaluationHistoryDependentQuestionnaire,
        )

        ehq = self.session.query(EvaluationHistoryDependentQuestionnaire).first()
        self.assertIsNotNone(ehq)
        ehq_id = ehq.id
        rv = self.client.get(
            "/api/q/evaluation_history_dependent_questionnaire/%i" % ehq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        response["self_identifies_autistic"] = False
        response["years_old_at_first_diagnosis"] = 12
        response["who_diagnosed"] = "healthTeam"
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/evaluation_history_dependent_questionnaire/%i" % ehq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/evaluation_history_dependent_questionnaire/%i" % ehq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["self_identifies_autistic"], False)
        self.assertEqual(response["years_old_at_first_diagnosis"], 12)
        self.assertEqual(response["who_diagnosed"], "healthTeam")
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_evaluation_history_dependent_questionnaire(self):
        ehq = self.construct_evaluation_history_dependent_questionnaire()
        ehq_id = ehq.id
        rv = self.client.get(
            "api/q/evaluation_history_dependent_questionnaire/%i" % ehq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/evaluation_history_dependent_questionnaire/%i" % ehq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/evaluation_history_dependent_questionnaire/%i" % ehq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_evaluation_history_dependent_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_guardian)
        headers = self.logged_in_headers(u)

        evaluation_history_dependent_questionnaire = {
            "self_identifies_autistic": True,
            "years_old_at_first_diagnosis": 5,
            "participant_id": p.id,
        }
        rv = self.client.post(
            "api/flow/dependent_intake/evaluation_history_dependent_questionnaire",
            data=self.jsonify(evaluation_history_dependent_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["self_identifies_autistic"], True)
        self.assertEqual(response["years_old_at_first_diagnosis"], 5)
        self.assertIsNotNone(response["id"])

    def test_evaluation_history_self_questionnaire_basics(self):
        self.construct_evaluation_history_self_questionnaire()
        from app.models import EvaluationHistorySelfQuestionnaire

        ehq = self.session.query(EvaluationHistorySelfQuestionnaire).first()
        self.assertIsNotNone(ehq)
        ehq_id = ehq.id
        rv = self.client.get(
            "/api/q/evaluation_history_self_questionnaire/%i" % ehq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], ehq_id)
        self.assertEqual(response["self_identifies_autistic"], ehq.self_identifies_autistic)
        self.assertEqual(response["years_old_at_first_diagnosis"], ehq.years_old_at_first_diagnosis)

    def test_modify_evaluation_history_self_questionnaire_basics(self):
        self.construct_evaluation_history_self_questionnaire()
        from app.models import EvaluationHistorySelfQuestionnaire

        ehq = self.session.query(EvaluationHistorySelfQuestionnaire).first()
        self.assertIsNotNone(ehq)
        ehq_id = ehq.id
        rv = self.client.get(
            "/api/q/evaluation_history_self_questionnaire/%i" % ehq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        response["self_identifies_autistic"] = False
        response["years_old_at_first_diagnosis"] = 12
        response["who_diagnosed"] = "healthTeam"
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/evaluation_history_self_questionnaire/%i" % ehq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/evaluation_history_self_questionnaire/%i" % ehq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["self_identifies_autistic"], False)
        self.assertEqual(response["years_old_at_first_diagnosis"], 12)
        self.assertEqual(response["who_diagnosed"], "healthTeam")
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_evaluation_history_self_questionnaire(self):
        ehq = self.construct_evaluation_history_self_questionnaire()
        ehq_id = ehq.id
        rv = self.client.get(
            "api/q/evaluation_history_self_questionnaire/%i" % ehq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/evaluation_history_self_questionnaire/%i" % ehq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/evaluation_history_self_questionnaire/%i" % ehq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_evaluation_history_self_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_guardian)
        headers = self.logged_in_headers(u)

        evaluation_history_self_questionnaire = {
            "self_identifies_autistic": True,
            "years_old_at_first_diagnosis": 5,
            "participant_id": p.id,
        }
        rv = self.client.post(
            "api/flow/self_intake/evaluation_history_self_questionnaire",
            data=self.jsonify(evaluation_history_self_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["self_identifies_autistic"], True)
        self.assertEqual(response["years_old_at_first_diagnosis"], 5)
        self.assertIsNotNone(response["id"])

    def test_home_dependent_questionnaire_basics(self):
        self.construct_home_dependent_questionnaire()
        from app.models import HomeDependentQuestionnaire

        hq = self.session.query(HomeDependentQuestionnaire).first()
        self.assertIsNotNone(hq)
        hq_id = hq.id
        rv = self.client.get(
            "/api/q/home_dependent_questionnaire/%i" % hq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], hq_id)
        self.assertEqual(response["participant_id"], hq.participant_id)
        self.assertEqual(response["user_id"], hq.user_id)
        self.assertEqual(response["dependent_living_situation"], hq.dependent_living_situation)
        self.assertEqual(response["struggle_to_afford"], hq.struggle_to_afford)
        self.assertEqual(len(response["housemates"]), len(hq.housemates))

    def test_modify_home_dependent_questionnaire_basics(self):
        self.construct_home_dependent_questionnaire()
        from app.models import HomeDependentQuestionnaire

        hq = self.session.query(HomeDependentQuestionnaire).first()
        self.assertIsNotNone(hq)
        hq_id = hq.id
        rv = self.client.get(
            "/api/q/home_dependent_questionnaire/%i" % hq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        from app.enums import Relationship

        response["participant_id"] = self.construct_participant(
            user=self.construct_user(), relationship=Relationship.dependent
        ).id
        response["dependent_living_situation"] = ["caregiver"]
        response["struggle_to_afford"] = True
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/home_dependent_questionnaire/%i" % hq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        self.construct_housemate(name="Debbie Danger", home_dependent_questionnaire=hq)
        rv = self.client.get(
            "/api/q/home_dependent_questionnaire/%i" % hq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["dependent_living_situation"], ["caregiver"])
        self.assertEqual(response["struggle_to_afford"], True)
        self.assertEqual(len(response["housemates"]), 2)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_home_dependent_questionnaire(self):
        hq = self.construct_home_dependent_questionnaire()
        hq_id = hq.id
        rv = self.client.get(
            "api/q/home_dependent_questionnaire/%i" % hq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/home_dependent_questionnaire/%i" % hq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/home_dependent_questionnaire/%i" % hq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_home_dependent_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        home_dependent_questionnaire = {
            "dependent_living_situation": ["family"],
            "struggle_to_afford": False,
            "participant_id": p.id,
        }
        rv = self.client.post(
            "api/flow/dependent_intake/home_dependent_questionnaire",
            data=self.jsonify(home_dependent_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["participant_id"], p.id)
        self.assertEqual(response["dependent_living_situation"], ["family"])
        self.assertEqual(response["struggle_to_afford"], False)
        self.assertIsNotNone(response["id"])

    def test_home_self_questionnaire_basics(self):
        self.construct_home_self_questionnaire()
        from app.models import HomeSelfQuestionnaire

        hq = self.session.query(HomeSelfQuestionnaire).first()
        self.assertIsNotNone(hq)
        hq_id = hq.id
        rv = self.client.get(
            "/api/q/home_self_questionnaire/%i" % hq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], hq_id)
        self.assertEqual(response["participant_id"], hq.participant_id)
        self.assertEqual(response["user_id"], hq.user_id)
        self.assertEqual(response["self_living_situation"], hq.self_living_situation)
        self.assertEqual(response["struggle_to_afford"], hq.struggle_to_afford)
        self.assertEqual(len(response["housemates"]), len(hq.housemates))

    def test_modify_home_self_questionnaire_basics(self):
        self.construct_home_self_questionnaire()
        from app.models import HomeSelfQuestionnaire

        hq = self.session.query(HomeSelfQuestionnaire).first()
        self.assertIsNotNone(hq)
        hq_id = hq.id
        rv = self.client.get(
            "/api/q/home_self_questionnaire/%i" % hq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        from app.enums import Relationship

        response["participant_id"] = self.construct_participant(
            user=self.construct_user(), relationship=Relationship.self_participant
        ).id
        response["self_living_situation"] = ["caregiver"]
        response["struggle_to_afford"] = True
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/home_self_questionnaire/%i" % hq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        self.construct_housemate(name="Debbie Danger", home_self_questionnaire=hq)
        rv = self.client.get(
            "/api/q/home_self_questionnaire/%i" % hq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["self_living_situation"], ["caregiver"])
        self.assertEqual(response["struggle_to_afford"], True)
        self.assertEqual(len(response["housemates"]), 2)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_home_self_questionnaire(self):
        hq = self.construct_home_self_questionnaire()
        hq_id = hq.id
        rv = self.client.get(
            "api/q/home_self_questionnaire/%i" % hq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/home_self_questionnaire/%i" % hq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/home_self_questionnaire/%i" % hq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_home_self_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        home_self_questionnaire = {
            "self_living_situation": ["family"],
            "struggle_to_afford": False,
            "participant_id": p.id,
        }
        rv = self.client.post(
            "api/flow/self_intake/home_self_questionnaire",
            data=self.jsonify(home_self_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["participant_id"], p.id)
        self.assertEqual(response["self_living_situation"], ["family"])
        self.assertEqual(response["struggle_to_afford"], False)
        self.assertIsNotNone(response["id"])

    def test_identification_questionnaire_basics(self):
        self.construct_identification_questionnaire()
        from app.models import IdentificationQuestionnaire

        iq = self.session.query(IdentificationQuestionnaire).first()
        self.assertIsNotNone(iq)
        iq_id = iq.id
        rv = self.client.get(
            "/api/q/identification_questionnaire/%i" % iq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], iq_id)
        self.assertEqual(response["first_name"], iq.first_name)
        self.assertEqual(response["nickname"], iq.nickname)
        self.assertEqual(response["birth_state"], iq.birth_state)

    def test_modify_identification_questionnaire_basics(self):
        self.construct_identification_questionnaire()
        from app.models import IdentificationQuestionnaire

        iq = self.session.query(IdentificationQuestionnaire).first()
        self.assertIsNotNone(iq)
        iq_id = iq.id
        rv = self.client.get(
            "/api/q/identification_questionnaire/%i" % iq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        response["first_name"] = "Helga"
        response["birth_city"] = "Staunton"
        response["is_first_name_preferred"] = True
        u2 = self.construct_user(email="rainbows@rainy.com")
        response["user_id"] = u2.id
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/identification_questionnaire/%i" % iq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        rv = self.client.get(
            "/api/q/identification_questionnaire/%i" % iq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["first_name"], "Helga")
        self.assertEqual(response["birth_city"], "Staunton")
        self.assertEqual(response["is_first_name_preferred"], True)
        self.assertEqual(response["user_id"], u2.id)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_identification_questionnaire(self):
        iq = self.construct_identification_questionnaire()
        iq_id = iq.id
        rv = self.client.get(
            "api/q/identification_questionnaire/%i" % iq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/identification_questionnaire/%i" % iq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/identification_questionnaire/%i" % iq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assertEqual(404, rv.status_code)

    def test_create_identification_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        identification_questionnaire = {"first_name": "Eloise", "middle_name": "Elora", "participant_id": p.id}
        rv = self.client.post(
            "api/flow/self_intake/identification_questionnaire",
            data=self.jsonify(identification_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["first_name"], "Eloise")
        self.assertEqual(response["middle_name"], "Elora")
        self.assertIsNotNone(response["id"])

    def test_supports_questionnaire_basics(self):
        self.construct_supports_questionnaire()
        from app.models import SupportsQuestionnaire

        sq = self.session.query(SupportsQuestionnaire).first()
        self.assertIsNotNone(sq)
        sq_id = sq.id
        rv = self.client.get(
            "/api/q/supports_questionnaire/%i" % sq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], sq_id)
        self.assertEqual(response["participant_id"], sq.participant_id)
        self.assertEqual(response["user_id"], sq.user_id)
        self.assertEqual(len(response["assistive_devices"]), len(sq.assistive_devices))
        self.assertEqual(len(response["medications"]), len(sq.medications))
        self.assertEqual(len(response["therapies"]), len(sq.therapies))

    def test_modify_supports_questionnaire_basics(self):
        self.construct_supports_questionnaire()
        from app.models import SupportsQuestionnaire

        sq = self.session.query(SupportsQuestionnaire).first()
        self.assertIsNotNone(sq)
        sq_id = sq.id
        rv = self.client.get(
            "/api/q/supports_questionnaire/%i" % sq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        response = rv.json
        from app.enums import Relationship

        response["participant_id"] = self.construct_participant(
            user=self.construct_user(), relationship=Relationship.self_participant
        ).id
        orig_date = response["last_updated"]
        rv = self.client.put(
            "/api/q/supports_questionnaire/%i" % sq_id,
            data=self.jsonify(response),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        self.construct_medication(name="Iocane Powder", supports_questionnaire=sq)
        self.construct_therapy(type="socialSkills", supports_questionnaire=sq)
        self.construct_alternative_augmentative(type="highTechAAC", supports_questionnaire=sq)
        self.construct_assistive_device(
            type_group="hearing",
            type="hearingAid",
            notes="Your ears you keep and I'll tell you why.",
            supports_questionnaire=sq,
        )
        rv = self.client.get(
            "/api/q/supports_questionnaire/%i" % sq_id,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(len(response["medications"]), 2)
        self.assertEqual(len(response["therapies"]), 2)
        self.assertEqual(len(response["alternative_augmentative"]), 2)
        self.assertEqual(len(response["assistive_devices"]), 2)
        self.assertNotEqual(orig_date, response["last_updated"])

    def test_delete_supports_questionnaire(self):
        sq = self.construct_supports_questionnaire()
        sq_id = sq.id
        rv = self.client.get(
            "api/q/supports_questionnaire/%i" % sq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/supports_questionnaire/%i" % sq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/supports_questionnaire/%i" % sq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assertEqual(404, rv.status_code)

    def test_create_supports_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        supports_questionnaire = {"participant_id": p.id}
        rv = self.client.post(
            "api/flow/self_intake/supports_questionnaire",
            data=self.jsonify(supports_questionnaire),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["participant_id"], p.id)
        self.assertIsNotNone(response["id"])

    def test_get_chain_steps(self):
        self.construct_chain_steps()
        from app.models import ChainStep

        chain_step = self.session.query(ChainStep).first()
        self.assertIsNotNone(chain_step)
        chain_step_id = chain_step.id
        chain_step_name = chain_step.name
        chain_step_instruction = chain_step.instruction
        rv = self.client.get(
            "/api/chain_step/%i" % chain_step_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], chain_step_id)
        self.assertEqual(response["name"], chain_step_name)
        self.assertEqual(response["instruction"], chain_step_instruction)

    def test_chain_session_questionnaire_basics(self):
        self.construct_chain_session_questionnaire()
        from app.models import ChainQuestionnaire

        cq = self.session.query(ChainQuestionnaire).first()
        self.assertIsNotNone(cq)
        cq_id = cq.id
        rv = self.client.get(
            "/api/q/chain_questionnaire/%i" % cq_id,
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["id"], cq_id)
        self.assertEqual(response["participant_id"], cq.participant_id)
        self.assertEqual(response["user_id"], cq.user_id)
        self.assertEqual(len(response["sessions"]), len(cq.sessions))
        self.assertEqual(len(response["sessions"][0]["step_attempts"]), len(cq.sessions[0].step_attempts))
        self.assertEqual(
            response["sessions"][0]["step_attempts"][0]["chain_step_id"], cq.sessions[0].step_attempts[0].chain_step_id
        )
        self.assertEqual(response["sessions"][0]["step_attempts"][0]["session_number"], len(cq.sessions))
        self.assertTrue("chain_step" in response["sessions"][0]["step_attempts"][0])
        self.assertIsNotNone(response["sessions"][0]["step_attempts"][0]["chain_step"]["name"])

    def test_modify_chain_session_questionnaire_basics(self):
        user = self.construct_user()
        from app.enums import Relationship

        participant = self.construct_participant(user=user, relationship=Relationship.self_participant)
        self.construct_chain_session_questionnaire()
        from app.models import ChainQuestionnaire

        cq = self.session.query(ChainQuestionnaire).first()
        self.assertIsNotNone(cq)
        cq_id = cq.id
        response_1 = self.client.get(
            "/api/q/chain_questionnaire/%i" % cq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        data_before = json.loads(response_1.get_data(as_text=True))

        data_before["user_id"] = user.id
        data_before["participant_id"] = participant.id
        now = datetime.datetime.now(tz=datetime.timezone.utc)
        later_1 = now + datetime.timedelta(minutes=1)
        later_2 = now + datetime.timedelta(minutes=2)
        later_3 = now + datetime.timedelta(minutes=3)
        data_before["sessions"] = [
            {
                "date": now.isoformat(),
                "completed": False,
                "session_type": "training",
                "step_attempts": [
                    {
                        "chain_step_id": 0,
                        "date": now.isoformat(),
                        "status": "focus",
                        "completed": False,
                        "was_prompted": True,
                        "prompt_level": "partial_physical",
                        "had_challenging_behavior": True,
                        "reason_step_incomplete": "challenging_behavior",
                        "challenging_behaviors": [
                            {"time": later_1.isoformat()},
                            {"time": later_2.isoformat()},
                            {"time": later_3.isoformat()},
                        ],
                    }
                ],
            }
        ]

        orig_date = data_before["last_updated"]
        response_2 = self.client.put(
            "/api/q/chain_questionnaire/%i" % cq_id,
            data=self.jsonify(data_before),
            content_type="application/json",
            follow_redirects=True,
            headers=self.logged_in_headers(),
        )
        self.assert_success(response_2)
        response_3 = self.client.get(
            "/api/q/chain_questionnaire/%i" % cq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assert_success(response_3)
        data_after = json.loads(response_3.get_data(as_text=True))
        self.assertNotEqual(orig_date, data_after["last_updated"])

        session_before = data_before["sessions"][0]
        session_after = data_after["sessions"][0]
        self.assertEqual(
            self._parse_date(session_before["date"]) - self._parse_date(session_after["date"]), datetime.timedelta(0)
        )
        self.assertTrue("step_attempts" in session_before)
        self.assertGreater(len(session_before["step_attempts"]), 0)

        self.assertTrue("step_attempts" in session_after)
        self.assertGreater(len(session_after["step_attempts"]), 0)

        step_attempt_before = session_before["step_attempts"][0]
        self.assertFalse("chain_step" in step_attempt_before)
        self.assertFalse("session_type" in step_attempt_before)
        self.assertFalse("was_focus_step" in step_attempt_before)
        self.assertFalse("target_prompt_level" in step_attempt_before)

        step_attempt_after = session_after["step_attempts"][0]
        self.assertTrue("chain_step" in step_attempt_after)
        self.assertTrue("session_type" in step_attempt_after)
        self.assertTrue("was_focus_step" in step_attempt_after)
        self.assertTrue("target_prompt_level" in step_attempt_after)

        self.assertEqual(step_attempt_before["prompt_level"], step_attempt_after["prompt_level"])
        from app.models import ChallengingBehavior

        all_cbs = self.session.query(ChallengingBehavior).all()
        self.assertEqual(len(all_cbs), 3)

        self.assertEqual(
            self._parse_date(step_attempt_before["date"]) - self._parse_date(step_attempt_after["date"]),
            datetime.timedelta(0),
        )
        self.assertEqual(
            self._parse_date(step_attempt_before["challenging_behaviors"][0]["time"])
            - self._parse_date(step_attempt_after["challenging_behaviors"][0]["time"]),
            datetime.timedelta(0),
        )

        # Export chain questionnaire?
        # Check for participant ID in chain session step
        rv_export_q = self.client.get(
            "/api/q/chain_questionnaire/export",
            follow_redirects=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv_export_q)
        wb_q = openpyxl.load_workbook(io.BytesIO(rv_export_q.data))
        ws_q = wb_q.active
        self.assertEqual("participant_id", ws_q["C1"].value)
        self.assertIsNotNone(ws_q["C2"].value)

        rv_export_session = self.client.get(
            "/api/q/chain_session/export",
            follow_redirects=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv_export_session)
        wb_session = openpyxl.load_workbook(io.BytesIO(rv_export_session.data))
        ws_session = wb_session.active
        self.assertEqual("participant_id", ws_session["C1"].value)
        self.assertIsNotNone(ws_session["C2"].value)

        rv_export_step = self.client.get(
            "/api/q/chain_session_step/export",
            follow_redirects=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv_export_step)
        wb_step = openpyxl.load_workbook(io.BytesIO(rv_export_step.data))
        ws_step = wb_step.active
        self.assertEqual("participant_id", ws_step["C1"].value)
        self.assertIsNotNone(ws_step["C2"].value)

    def test_delete_chain_session_questionnaire(self):
        sq = self.construct_chain_session_questionnaire()
        sq_id = sq.id
        rv = self.client.get(
            "api/q/chain_questionnaire/%i" % sq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assert_success(rv)

        rv = self.client.delete(
            "api/q/chain_questionnaire/%i" % sq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assert_success(rv)

        rv = self.client.get(
            "api/q/chain_questionnaire/%i" % sq_id, content_type="application/json", headers=self.logged_in_headers()
        )
        self.assertEqual(404, rv.status_code)

    def test_create_chain_session_questionnaire(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        cq = {"participant_id": p.id}
        rv = self.client.post(
            "api/flow/skillstar/chain_questionnaire",
            data=self.jsonify(cq),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(response["participant_id"], p.id)
        self.assertIsNotNone(response["id"])

    def test_flow_endpoint(self):
        # It should be possible to get a list of available flows
        rv = self.client.get("api/flow", content_type="application/json", headers=self.logged_in_headers())
        self.assertEqual(200, rv.status_code)
        response = rv.json
        self.assertIsNotNone(response)
        self.assertTrue(len(response) > 0)

    def test_intake_flows_endpoint(self):
        # Are the basics correct about the existing intake flows?
        rv = self.client.get("api/flow", content_type="application/json", headers=self.logged_in_headers())
        self.assertEqual(200, rv.status_code)
        response = rv.json
        for i in response:
            if i["name"] == "self_intake":
                self.assertEqual(len(i["steps"]), 10)
                self.assertEqual(i["steps"][8]["name"], "employment_questionnaire")
                self.assertEqual(i["steps"][8]["label"], "Employment")
            if i["name"] == "dependent_intake":
                self.assertEqual(len(i["steps"]), 9)
                self.assertEqual(i["steps"][5]["name"], "developmental_questionnaire")
                self.assertEqual(i["steps"][5]["label"], "Birth and Developmental History")
            if i["name"] == "guardian_intake":
                self.assertEqual(len(i["steps"]), 3)
                self.assertEqual(i["steps"][1]["name"], "contact_questionnaire")
                self.assertEqual(i["steps"][1]["label"], "Contact Information")

    def test_self_intake_flow_with_user(self):
        u = self.construct_user()
        from app.enums import Relationship

        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)
        rv = self.client.get("api/flow/self_intake/%i" % p.id, content_type="application/json", headers=headers)
        self.assertEqual(200, rv.status_code)
        response = rv.json
        self.assertIsNotNone(response)
        self.assertEqual("self_intake", response["name"])
        self.assertIsNotNone(response["steps"])
        self.assertTrue(len(response["steps"]) > 0)
        self.assertEqual("identification_questionnaire", response["steps"][0]["name"])
        self.assertEqual(ExportService.TYPE_IDENTIFYING, response["steps"][0]["type"])
        from app.models import Step

        self.assertEqual(Step.STATUS_INCOMPLETE, response["steps"][0]["status"])

        cq = self.get_identification_questionnaire(p.id)
        rv = self.client.post(
            "api/flow/self_intake/identification_questionnaire",
            data=self.jsonify(cq),
            content_type="application/json",
            follow_redirects=True,
            headers=headers,
        )

        rv = self.client.get("api/flow/self_intake/%i" % p.id, content_type="application/json", headers=headers)
        response = rv.json
        self.assertEqual("identification_questionnaire", response["steps"][0]["name"])
        self.assertEqual(Step.STATUS_COMPLETE, response["steps"][0]["status"])
        self.assertIsNotNone(response["steps"][0]["date_completed"])

    def test_questionnaire_meta_is_relation_specific(self):
        rv = self.client.get(
            "/api/flow/self_intake/identification_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        intro = self.get_field_from_response(response, "intro")
        self.assertIsNotNone(intro["template_options"]["description"])
        self.assertEqual(
            intro["template_options"]["description"],
            "Please answer the following questions about yourself (* indicates required response):",
        )

        birth_city = self.get_field_from_response(response, "birth_city")
        self.assertIsNotNone(birth_city)
        self.assertIsNotNone(birth_city["template_options"])
        self.assertIsNotNone(birth_city["template_options"]["label"])
        self.assertEqual(birth_city["template_options"]["label"], "Your city/municipality of birth")

        rv = self.client.get(
            "/api/flow/dependent_intake/identification_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        birth_city = self.get_field_from_response(response, "birth_city")
        self.assertIsNotNone(birth_city["template_options"]["label"])
        self.assertEqual(birth_city["template_options"]["label"], "Your child's city/municipality of birth")

    def test_questionnaire_meta_has_relation_required_fields(self):
        rv = self.client.get(
            "/api/flow/dependent_intake/identification_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        relationship = self.get_field_from_response(response, "relationship")
        self.assertIsNotNone(relationship)

        # Convert Participant to a dependant
        rv = self.client.get(
            "/api/flow/self_intake/identification_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        relationship = self.get_field_from_response(response, "relationship")
        self.assertIsNone(relationship)

    def test_meta_contains_table_details(self):
        rv = self.client.get(
            "/api/flow/dependent_intake/identification_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual("identifying", response["table"]["question_type"])
        self.assertEqual("Identification", response["table"]["label"])

    def test_meta_field_groups_contain_their_fields(self):
        rv = self.client.get(
            "/api/flow/self_intake/home_self_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self_living = self.get_field_from_response(response, "self_living")
        self.assertEqual("self_living_situation", self_living["fieldGroup"][0]["name"])

    def test_support_meta_contain_their_fields(self):
        rv = self.client.get(
            "/api/flow/self_intake/supports_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        assistive_devices = self.get_field_from_response(response, "assistive_devices")
        self.assertIsNotNone(assistive_devices["fieldArray"]["fieldGroup"][0])
        self.assertEqual("type", assistive_devices["fieldArray"]["fieldGroup"][0]["name"])

    def test_evaluation_history_dependent_meta_contain_their_fields(self):
        rv = self.client.get(
            "/api/flow/dependent_intake/evaluation_history_dependent_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(len(response["fields"]), 10)

    def test_evaluation_history_self_meta_contain_their_fields(self):
        rv = self.client.get(
            "/api/flow/self_intake/evaluation_history_self_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(len(response["fields"]), 10)

    def test_education_dependent_meta_contain_their_fields(self):
        rv = self.client.get(
            "/api/flow/dependent_intake/education_dependent_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(len(response["fields"]), 5)

    def test_education_self_meta_contain_their_fields(self):
        rv = self.client.get(
            "/api/flow/self_intake/education_self_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(len(response["fields"]), 5)

    def test_meta_fields_are_ordered(self):
        rv = self.client.get(
            "/api/flow/self_intake/education_self_questionnaire/meta",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(1, response["fields"][0]["display_order"])
        self.assertEqual(2, response["fields"][1]["display_order"])
        self.assertEqual(3, response["fields"][2]["display_order"])

        self.assertEqual(6.1, response["fields"][4]["fieldGroup"][0]["display_order"])
        self.assertEqual(6.2, response["fields"][4]["fieldGroup"][1]["display_order"])

    def test_questionnaire_info_list_basics(self):
        rv = self.client.get("/api/q", follow_redirects=True, content_type="application/json")
        self.assert_success(rv)
        response = rv.json
        self.assertEqual("Contact", response[4]["display_name"])
        self.assertEqual("Employment", response[11]["display_name"])
        self.assertEqual("Professional Profile", response[17]["display_name"])
        self.assertEqual("Supports", response[19]["display_name"])
        self.assertEqual(4, len(response[19]["sub_tables"]))
        self.assertEqual(20, len(response))

    def test_questionnaire_list_meta_basics(self):
        rv = self.client.get(
            "/api/q/education_self_questionnaire/meta", follow_redirects=True, content_type="application/json"
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(1, len(list(filter(lambda field: field["name"] == "id", response["fields"]))))
        self.assertEqual(1, len(list(filter(lambda field: field["name"] == "user_id", response["fields"]))))
        self.assertEqual(1, len(list(filter(lambda field: field["name"] == "school_type", response["fields"]))))
        self.assertEqual(
            1, len(list(filter(lambda field: field["name"] == "school_services_other", response["fields"])))
        )
        self.assertEqual(13, len(response["fields"]))

    def test_questionnaire_list_basics(self):
        q = self.construct_current_behaviors_dependent_questionnaire()
        rv = self.client.get(
            "/api/q/current_behaviors_dependent_questionnaire",
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        response = rv.json
        self.assertEqual(["math", "writing"], response[0]["academic_difficulty_areas"])
        self.assertEqual("fluent", response[0]["dependent_verbal_ability"])
        self.assertEqual(q.id, response[0]["id"])

    def test_non_admin_cannot_view_questionnaire_list(self):
        user = self.construct_user(email="regularUser@user.com")
        admin = self.construct_admin_user(email="adminUser@user.com")
        self.construct_contact_questionnaire()
        rv = self.client.get("/api/q/contact_questionnaire", content_type="application/json")
        self.assertEqual(401, rv.status_code)
        rv = self.client.get(
            "/api/q/contact_questionnaire", content_type="application/json", headers=self.logged_in_headers(user=user)
        )
        self.assertEqual(403, rv.status_code)
        rv = self.client.get(
            "/api/q/contact_questionnaire", content_type="application/json", headers=self.logged_in_headers(user=admin)
        )
        self.assert_success(rv)

    def test_export_single_questionnaire(self):
        self.construct_current_behaviors_dependent_questionnaire()
        rv = self.client.get(
            "/api/q/current_behaviors_dependent_questionnaire/export",
            follow_redirects=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        wb = openpyxl.load_workbook(io.BytesIO(rv.data))
        ws = wb.active
        self.assertEqual("id", ws["A1"].value)
        self.assertEqual("last_updated", ws["B1"].value)
        self.assertEqual("participant_id", ws["C1"].value)
        self.assertEqual("user_id", ws["D1"].value)
        self.assertEqual("time_on_task_ms", ws["E1"].value)
        self.assertEqual("dependent_verbal_ability", ws["F1"].value)
        self.assertEqual("concerning_behaviors", ws["G1"].value)
        self.assertEqual("hoarding, ", ws["G2"].value)
        self.assertEqual("concerning_behaviors_other", ws["H1"].value)
        self.assertEqual("has_academic_difficulties", ws["I1"].value)
        self.assertEqual("academic_difficulty_areas", ws["J1"].value)
        self.assertEqual("math, writing, ", ws["J2"].value)
        self.assertEqual("academic_difficulty_other", ws["K1"].value)
        self.assertEqual(11, ws.max_column)
        self.assertEqual(2, ws.max_row)

    def test_export_all_questionnaires(self):
        self.construct_all_questionnaires()
        rv = self.client.get(
            "/api/q/all/export",
            follow_redirects=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        wb = openpyxl.load_workbook(io.BytesIO(rv.data))
        ws = wb.active
        self.assertEqual(2, ws.max_row)
        self.assertEqual(26, len(wb.worksheets))
        worksheet_titles = [ws.title for ws in wb.worksheets]
        self.assertEqual(
            [
                "Alternative Augmentative",
                "Assistive Device",
                "Chain",
                "Chain Session",
                "Chain Session Step",
                "Challenging Behavior",
                "Clinical Diagnoses",
                "Contact",
                "Current Behaviors Dependent",
                "Current Behaviors Self",
                "Demographics",
                "Developmental",
                "Education Dependent",
                "Education Self",
                "Employment",
                "Evaluation History Dependent",
                "Evaluation History Self",
                "Home Dependent",
                "Home Self",
                "Housemate",
                "Identification",
                "Medication",
                "Professional Profile",
                "Registration",
                "Supports",
                "Therapy",
            ],
            worksheet_titles,
        )

    def test_export_questionnaires_by_user(self):
        u1 = self.construct_user(email="1@sartography.com")
        u2 = self.construct_user(email="2@sartography.com")
        self.construct_all_questionnaires(u1)
        self.construct_all_questionnaires(u2)
        rv = self.client.get(
            "/api/q/all/export/user/%i" % u1.id,
            follow_redirects=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        wb = openpyxl.load_workbook(io.BytesIO(rv.data))
        self.assertEqual(26, len(wb.worksheets))
        self.assertEqual(2, wb["Contact"].max_row)
        self.assertEqual("user_id", wb["Contact"]["E1"].value)
        self.assertEqual(u1.id, wb["Contact"]["E2"].value)
        rv = self.client.get(
            "/api/q/all/export/user/%i" % u2.id,
            follow_redirects=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=self.logged_in_headers(),
        )
        self.assert_success(rv)
        wb = openpyxl.load_workbook(io.BytesIO(rv.data))
        self.assertEqual(2, wb["Contact"].max_row)
        self.assertEqual("user_id", wb["Contact"]["E1"].value)
        self.assertEqual(u2.id, wb["Contact"]["E2"].value)

    def _parse_date(self, date_str):
        return parser.parse(date_str).replace(tzinfo=None)

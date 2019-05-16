# Set environment variable to testing before loading.
# IMPORTANT - Environment must be loaded before app, models, etc....
import os

os.environ["APP_CONFIG_FILE"] = '../config/testing.py'

from app.model.flow import Step
from app.model.step_log import StepLog
from app.question_service import QuestionService
import re
import json
import base64
import quopri
import random
import string
import unittest
import openpyxl
import io
from app import db, app, elastic_index
from app.model.user import User, Role
from app.model.study import Study, Status
from app.email_service import TEST_MESSAGES
from app.model.category import Category
from app.model.event import Event
from app.model.location import Location
from app.model.resource import StarResource
from app.model.email_log import EmailLog
from app.model.organization import Organization
from app.model.investigator import Investigator
from app.model.participant import Participant, Relationship
from app.model.study_investigator import StudyInvestigator
from app.model.study_category import StudyCategory
from app.model.resource_category import ResourceCategory
from app.model.questionnaires.assistive_device import AssistiveDevice
from app.model.questionnaires.alternative_augmentative import AlternativeAugmentative
from app.model.questionnaires.clinical_diagnoses_questionnaire import ClinicalDiagnosesQuestionnaire
from app.model.questionnaires.contact_questionnaire import ContactQuestionnaire
from app.model.questionnaires.current_behaviors_dependent_questionnaire import CurrentBehaviorsDependentQuestionnaire
from app.model.questionnaires.current_behaviors_self_questionnaire import CurrentBehaviorsSelfQuestionnaire
from app.model.questionnaires.demographics_questionnaire import DemographicsQuestionnaire
from app.model.questionnaires.developmental_questionnaire import DevelopmentalQuestionnaire
from app.model.questionnaires.education_dependent_questionnaire import EducationDependentQuestionnaire
from app.model.questionnaires.education_self_questionnaire import EducationSelfQuestionnaire
from app.model.questionnaires.employment_questionnaire import EmploymentQuestionnaire
from app.model.questionnaires.evaluation_history_dependent_questionnaire import EvaluationHistoryDependentQuestionnaire
from app.model.questionnaires.evaluation_history_self_questionnaire import EvaluationHistorySelfQuestionnaire
from app.model.questionnaires.home_dependent_questionnaire import HomeDependentQuestionnaire
from app.model.questionnaires.home_self_questionnaire import HomeSelfQuestionnaire
from app.model.questionnaires.housemate import Housemate
from app.model.questionnaires.identification_questionnaire import IdentificationQuestionnaire
from app.model.questionnaires.professional_profile_questionnaire import ProfessionalProfileQuestionnaire
from app.model.questionnaires.medication import Medication
from app.model.questionnaires.supports_questionnaire import SupportsQuestionnaire
from app.model.questionnaires.therapy import Therapy


class TestCase(unittest.TestCase):

    def setUp(self):
        self.ctx = app.test_request_context()
        self.app = app.test_client()
        db.session.remove()
        db.drop_all()
        db.create_all()
        self.ctx.push()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        # elastic_index.clear()
        self.ctx.pop()

    def assertSuccess(self, rv):
        try:
            data = json.loads(rv.get_data(as_text=True))
            self.assertTrue(rv.status_code >= 200 and rv.status_code < 300,
                            "BAD Response: %i. \n %s" %
                            (rv.status_code, json.dumps(data)))
        except:
            self.assertTrue(rv.status_code >= 200 and rv.status_code < 300,
                            "BAD Response: %i." % rv.status_code)

    def randomString(self):
        char_set = string.ascii_uppercase + string.digits
        return ''.join(random.sample(char_set * 6, 6))

    def test_base_endpoint(self):
        rv = self.app.get('/',
                          follow_redirects=True,
                          content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))

        endpoints = [
            ('api.categorybyresourceendpoint', '/api/resource/<resource_id>/category'),
            ('api.categorybystudyendpoint', '/api/study/<study_id>/category'),
            ('api.categoryendpoint', '/api/category/<id>'),
            ('api.categorylistendpoint', '/api/category'),
            ('api.questionnaireendpoint', '/api/q/<name>/<id>'),
            ('api.organizationendpoint', '/api/organization/<id>'),
            ('api.organizationlistendpoint', '/api/organization'),
            ('api.resourcebycategoryendpoint', '/api/category/<category_id>/resource'),
            ('api.resourcecategoryendpoint', '/api/resource_category/<id>'),
            ('api.resourcecategorylistendpoint', '/api/resource_category'),
            ('api.resourceendpoint', '/api/resource/<id>'),
            ('api.resourcelistendpoint', '/api/resource'),
            ('api.rootcategorylistendpoint', '/api/category/root'),
            ('api.sessionendpoint', '/api/session'),
            ('api.studybycategoryendpoint', '/api/category/<category_id>/study'),
            ('api.studycategoryendpoint', '/api/study_category/<id>'),
            ('api.studycategorylistendpoint', '/api/study_category'),
            ('api.studyendpoint', '/api/study/<id>'),
            ('api.studylistendpoint', '/api/study'),
            ('api.userendpoint', '/api/user/<id>'),
            ('api.userlistendpoint', '/api/user'),
            ('auth.forgot_password', '/api/forgot_password'),
            ('auth.login_password', '/api/login_password'),
            ('auth.reset_password', '/api/reset_password'),
        ]

        for endpoint in endpoints:
            self.assertEqual(response[endpoint[0]], endpoint[1])

    def get_field_from_response(self, response, name):
        for field in response['fields']:
            if field["name"] == name:
                return field

    def construct_event(self, title="A+ Event", description="A delightful event destined to create rejoicing",
                           street_address1="123 Some Pl", street_address2="Apt. 45",
                           city="Stauntonville", state="QX", zip="99775", phone="555-555-5555",
                           website="http://stardrive.org"):

        event = Event(title=title, description=description, street_address1=street_address1, street_address2=street_address2, city=city,
                                state=state, zip=zip, phone=phone, website=website)
        event.organization_id = self.construct_organization().id
        db.session.add(event)
        db.session.commit()

        db_event = db.session.query(Event).filter_by(title=event.title).first()
        self.assertEqual(db_event.website, event.website)
        elastic_index.add_document(db_event, 'Event')
        return db_event

    def construct_location(self, title="A+ location", description="A delightful location destined to create rejoicing",
                           street_address1="123 Some Pl", street_address2="Apt. 45",
                           city="Stauntonville", state="QX", zip="99775", phone="555-555-5555",
                           website="http://stardrive.org"):

        location = Location(title=title, description=description, street_address1=street_address1, street_address2=street_address2, city=city,
                                state=state, zip=zip,phone=phone, website=website)
        location.organization_id = self.construct_organization().id
        db.session.add(location)
        db.session.commit()

        db_location = db.session.query(Location).filter_by(title=location.title).first()
        self.assertEqual(db_location.website, location.website)
        elastic_index.add_document(db_location, 'Location')
        return db_location

    def construct_resource(self, title="A+ Resource", description="A delightful Resource destined to create rejoicing",
                           phone="555-555-5555", website="http://stardrive.org"):

        resource = StarResource(title=title, description=description, phone=phone, website=website)
        resource.organization_id = self.construct_organization().id
        db.session.add(resource)
        db.session.commit()

        db_resource = db.session.query(StarResource).filter_by(title=resource.title).first()
        self.assertEqual(db_resource.website, resource.website)
        elastic_index.add_document(db_resource, 'Resource')
        return db_resource

    def construct_study(self, title="Fantastic Study", description="A study that will go down in history",
                        participant_description="Even your pet hamster could benefit from participating in this study",
                        benefit_description="You can expect to have your own rainbow following you around afterwards"):

        study = Study(title=title, description=description, participant_description=participant_description,
                      benefit_description=benefit_description, status=Status.currently_enrolling)
        study.organization_id = self.construct_organization().id
        db.session.add(study)
        db.session.commit()

        db_study = db.session.query(Study).filter_by(title=study.title).first()
        self.assertEqual(db_study.description, study.description)
        return db_study

    def construct_organization(self, name="Staunton Makerspace",
                               description="A place full of surprise, delight, and amazing people. And tools. Lots of exciting tools."):

        organization = Organization(name=name, description=description)
        db.session.add(organization)
        db.session.commit()

        db_org = db.session.query(Organization).filter_by(name=organization.name).first()
        self.assertEqual(db_org.description, organization.description)
        return db_org

    def construct_category(self, name="Ultimakers", parent=None):

        category = Category(name=name)
        if parent is not None:
            category.parent = parent
        db.session.add(category)
        db.session.commit()

        db_category = db.session.query(Category).filter_by(name=category.name).first()
        self.assertIsNotNone(db_category.id)
        return db_category

    def construct_investigator(self, name="Judith Wonder", title="Ph.D., Assistant Professor of Mereology"):

        investigator = Investigator(name=name, title=title)
        investigator.organization_id = self.construct_organization().id
        db.session.add(investigator)
        db.session.commit()

        db_inv = db.session.query(Investigator).filter_by(name=investigator.name).first()
        self.assertEqual(db_inv.title, investigator.title)
        return db_inv

    def construct_user(self, email="stan@staunton.com"):

        user = User(email=email, role=Role.user)
        db.session.add(user)
        db.session.commit()

        db_user = db.session.query(User).filter_by(email=user.email).first()
        self.assertEqual(db_user.email, user.email)
        return db_user

    def construct_admin_user(self, email="rmond@virginia.gov"):

        user = User(email=email, role=Role.admin)
        db.session.add(user)
        db.session.commit()

        db_user = db.session.query(User).filter_by(email=user.email).first()
        self.assertEqual(db_user.role, user.role)
        return db_user

    def construct_participant(self, user, relationship):

        participant = Participant(user=user, relationship=relationship)
        db.session.add(participant)
        db.session.commit()

        db_participant = db.session.query(Participant).filter_by(id=participant.id).first()
        self.assertEqual(db_participant.relationship, participant.relationship)
        return db_participant

    def construct_assistive_device(self, type_group='mobility', type='prosthetic', timeframe='current',
                                   notes='I love my new leg!', supports_questionnaire=None):

        ad = AssistiveDevice(type=type, timeframe=timeframe, notes=notes)
        if supports_questionnaire is not None:
            ad.supports_questionnaire_id = supports_questionnaire.id

        db.session.add(ad)
        db.session.commit()

        db_ad = db.session.query(AssistiveDevice).filter_by(last_updated=ad.last_updated).first()
        self.assertEqual(db_ad.notes, ad.notes)
        self.assertEqual(db_ad.type_group, ad.type_group)
        self.assertEqual(db_ad.type, ad.type)
        return db_ad

    def construct_alternative_augmentative(self, type='lowTechAAC', timeframe='current', notes='We use pen and paper', supports_questionnaire=None):

        aac = AlternativeAugmentative(type=type, timeframe=timeframe, notes=notes)
        if supports_questionnaire is not None:
            aac.supports_questionnaire_id = supports_questionnaire.id

        db.session.add(aac)
        db.session.commit()

        db_aac = db.session.query(AlternativeAugmentative).filter_by(last_updated=aac.last_updated).first()
        self.assertEqual(db_aac.notes, aac.notes)
        return db_aac

    def construct_clinical_diagnoses_questionnaire(self, developmental=['speechLanguage'], mental_health=['ocd'],
                                                   medical=['insomnia'], genetic=['corneliaDeLange'], participant=None,
                                                   user=None):

        cdq = ClinicalDiagnosesQuestionnaire(developmental=developmental, mental_health=mental_health, medical=medical,
                                             genetic=genetic)
        if user is None:
            u = self.construct_user(email='clinic@questionnaire.com')
            cdq.user_id = u.id
        else:
            u = user
            cdq.user_id = u.id

        if participant is None:
            cdq.participant_id = self.construct_participant(user=u, relationship=Relationship.dependent).id
        else:
            cdq.participant_id = participant.id

        db.session.add(cdq)
        db.session.commit()

        db_cdq = db.session.query(ClinicalDiagnosesQuestionnaire).filter_by(user_id=cdq.user_id).first()
        self.assertEqual(db_cdq.participant_id, cdq.participant_id)
        return db_cdq

    def construct_contact_questionnaire(self, phone="123-456-7890", zip=55678, marketing_channel="Zine Ad",
                                        participant=None, user=None):

        cq = ContactQuestionnaire(phone=phone, zip=zip,
                                  marketing_channel=marketing_channel)
        if user is None:
            u = self.construct_user(email='contact@questionnaire.com')
            cq.user_id = u.id
        else:
            u = user
            cq.user_id = u.id

        if participant is None:
            cq.participant_id = self.construct_participant(user=u, relationship=Relationship.dependent).id
        else:
            cq.participant_id = participant.id

        db.session.add(cq)
        db.session.commit()

        db_cq = db.session.query(ContactQuestionnaire).filter_by(zip=cq.zip).first()
        self.assertEqual(db_cq.phone, cq.phone)
        return db_cq

    def construct_current_behaviors_dependent_questionnaire(self, dependent_verbal_ability='fluent',
                                                            concerning_behaviors=['hoarding'],
                                                            has_academic_difficulties=True,
                                                            academic_difficulty_areas=['math', 'writing'],
                                                            participant=None, user=None):

        cb = CurrentBehaviorsDependentQuestionnaire(dependent_verbal_ability=dependent_verbal_ability,
                                                    concerning_behaviors=concerning_behaviors,
                                                    has_academic_difficulties=has_academic_difficulties,
                                                    academic_difficulty_areas=academic_difficulty_areas)
        if user is None:
            u = self.construct_user(email='cbd@questionnaire.com')
            cb.user_id = u.id
        else:
            u = user
            cb.user_id = u.id

        if participant is None:
            p = self.construct_participant(user=u, relationship=Relationship.dependent)
            cb.participant_id = p.id
        else:
            p = participant
            cb.participant_id = p.id

        db.session.add(cb)
        db.session.commit()

        db_cbd = db.session.query(CurrentBehaviorsDependentQuestionnaire).filter_by(
            participant_id=cb.participant_id).first()
        self.assertEqual(db_cbd.concerning_behaviors, cb.concerning_behaviors)
        return db_cbd

    def construct_current_behaviors_self_questionnaire(self, self_verbal_ability='verbal',
                                                       has_academic_difficulties=True, academic_difficulty_areas='math',
                                                       participant=None, user=None):

        cb = CurrentBehaviorsSelfQuestionnaire(self_verbal_ability=self_verbal_ability,
                                               has_academic_difficulties=has_academic_difficulties,
                                               academic_difficulty_areas=academic_difficulty_areas)
        if user is None:
            u = self.construct_user(email='cbs@questionnaire.com')
            cb.user_id = u.id
        else:
            u = user
            cb.user_id = u.id

        if participant is None:
            p = self.construct_participant(user=u, relationship=Relationship.self_participant)
            cb.participant_id = p.id
        else:
            p = participant
            cb.participant_id = p.id

        db.session.add(cb)
        db.session.commit()

        db_cb = db.session.query(CurrentBehaviorsSelfQuestionnaire).filter_by(participant_id=cb.participant_id).first()
        self.assertEqual(db_cb.academic_difficulty_areas, cb.academic_difficulty_areas)
        return db_cb

    def construct_demographics_questionnaire(self, birth_sex="intersex", gender_identity="intersex",
                                             race_ethnicity="raceBlack", participant=None, user=None):

        dq = DemographicsQuestionnaire(birth_sex=birth_sex, gender_identity=gender_identity,
                                       race_ethnicity=race_ethnicity)
        if user is None:
            u = self.construct_user(email='demograph@questionnaire.com')
            dq.user_id = u.id
        else:
            u = user
            dq.user_id = u.id

        if participant is None:
            dq.participant_id = self.construct_participant(user=u, relationship=Relationship.self_participant).id
        else:
            dq.participant_id = participant.id

        db.session.add(dq)
        db.session.commit()

        db_dq = db.session.query(DemographicsQuestionnaire).filter_by(birth_sex=dq.birth_sex).first()
        self.assertEqual(db_dq.gender_identity, dq.gender_identity)
        return db_dq

    def construct_developmental_questionnaire(self, had_birth_complications=False, when_motor_milestones='delayed',
                                              when_language_milestones='early', when_toileting_milestones='notYet',
                                              participant=None, user=None):

        dq = DevelopmentalQuestionnaire(had_birth_complications=had_birth_complications,
                                        when_motor_milestones=when_motor_milestones,
                                        when_language_milestones=when_language_milestones,
                                        when_toileting_milestones=when_toileting_milestones)
        if user is None:
            u = self.construct_user(email='develop@questionnaire.com')
            dq.user_id = u.id
        else:
            u = user
            dq.user_id = u.id

        if participant is None:
            dq.participant_id = self.construct_participant(user=u, relationship=Relationship.dependent).id
        else:
            dq.participant_id = participant.id

        db.session.add(dq)
        db.session.commit()

        db_dq = db.session.query(DevelopmentalQuestionnaire).filter_by(participant_id=dq.participant_id).first()
        self.assertEqual(db_dq.when_language_milestones, dq.when_language_milestones)
        return db_dq

    def construct_education_dependent_questionnaire(self, attends_school=True, school_name='Harvard',
                                                    school_type='privateSchool',
                                                    dependent_placement='graduate', participant=None, user=None):

        eq = EducationDependentQuestionnaire(attends_school=attends_school, school_name=school_name,
                                             school_type=school_type,
                                             dependent_placement=dependent_placement)
        if user is None:
            u = self.construct_user(email='edudep@questionnaire.com')
            eq.user_id = u.id
        else:
            u = user
            eq.user_id = u.id

        if participant is None:
            p = self.construct_participant(user=u, relationship=Relationship.dependent)
            eq.participant_id = p.id
        else:
            p = participant
            eq.participant_id = p.id

        db.session.add(eq)
        db.session.commit()

        db_eq = db.session.query(EducationDependentQuestionnaire).filter_by(participant_id=eq.participant_id).first()
        self.assertEqual(db_eq.school_name, eq.school_name)
        return db_eq

    def construct_education_self_questionnaire(self, attends_school=True, school_name='Harvard',
                                               school_type='privateSchool',
                                               self_placement='graduate', participant=None, user=None):

        eq = EducationSelfQuestionnaire(attends_school=attends_school, school_name=school_name, school_type=school_type,
                                        self_placement=self_placement)

        if user is None:
            u = self.construct_user(email='eduself@questionnaire.com')
            eq.user_id = u.id
        else:
            u = user
            eq.user_id = u.id

        if participant is None:
            p = self.construct_participant(user=u, relationship=Relationship.self_participant)
            eq.participant_id = p.id
        else:
            p = participant
            eq.participant_id = p.id

        db.session.add(eq)
        db.session.commit()

        db_eq = db.session.query(EducationSelfQuestionnaire).filter_by(participant_id=eq.participant_id).first()
        self.assertEqual(db_eq.school_name, eq.school_name)
        return db_eq

    def construct_employment_questionnaire(self, is_currently_employed=True, employment_capacity='fullTime',
                                           has_employment_support=False, participant=None, user=None):

        eq = EmploymentQuestionnaire(is_currently_employed=is_currently_employed,
                                     employment_capacity=employment_capacity,
                                     has_employment_support=has_employment_support)
        if user is None:
            u = self.construct_user(email='employ@questionnaire.com')
            eq.user_id = u.id
        else:
            u = user
            eq.user_id = u.id

        if participant is None:
            eq.participant_id = self.construct_participant(user=u, relationship=Relationship.self_participant).id
        else:
            eq.participant_id = participant.id

        db.session.add(eq)
        db.session.commit()

        db_eq = db.session.query(EmploymentQuestionnaire).filter_by(participant_id=eq.participant_id).first()
        self.assertEqual(db_eq.employment_capacity, eq.employment_capacity)
        return db_eq

    def construct_evaluation_history_dependent_questionnaire(self, self_identifies_autistic=True,
                                                             has_autism_diagnosis=True,
                                                             years_old_at_first_diagnosis=7,
                                                             who_diagnosed="pediatrician",
                                                             participant=None, user=None):

        ehq = EvaluationHistoryDependentQuestionnaire(self_identifies_autistic=self_identifies_autistic,
                                                      has_autism_diagnosis=has_autism_diagnosis,
                                                      years_old_at_first_diagnosis=years_old_at_first_diagnosis,
                                                      who_diagnosed=who_diagnosed)
        if user is None:
            u = self.construct_user(email='evaldep@questionnaire.com')
            ehq.user_id = u.id
        else:
            u = user
            ehq.user_id = u.id

        if participant is None:
            p = self.construct_participant(user=u, relationship=Relationship.dependent)
            ehq.participant_id = p.id
        else:
            p = participant
            ehq.participant_id = p.id

        db.session.add(ehq)
        db.session.commit()

        db_ehq = db.session.query(EvaluationHistoryDependentQuestionnaire).filter_by(
            participant_id=ehq.participant_id).first()
        self.assertEqual(db_ehq.who_diagnosed, ehq.who_diagnosed)
        return db_ehq

    def construct_evaluation_history_self_questionnaire(self, self_identifies_autistic=True, has_autism_diagnosis=True,
                                                        years_old_at_first_diagnosis=7, who_diagnosed="pediatrician",
                                                        participant=None, user=None):

        ehq = EvaluationHistorySelfQuestionnaire(self_identifies_autistic=self_identifies_autistic,
                                                 has_autism_diagnosis=has_autism_diagnosis,
                                                 years_old_at_first_diagnosis=years_old_at_first_diagnosis,
                                                 who_diagnosed=who_diagnosed)
        if user is None:
            u = self.construct_user(email='evalself@questionnaire.com')
            ehq.user_id = u.id
        else:
            u = user
            ehq.user_id = u.id

        if participant is None:
            p = self.construct_participant(user=u, relationship=Relationship.self_participant)
            ehq.participant_id = p.id
        else:
            p = participant
            ehq.participant_id = p.id

        db.session.add(ehq)
        db.session.commit()

        db_ehq = db.session.query(EvaluationHistorySelfQuestionnaire).filter_by(
            participant_id=ehq.participant_id).first()
        self.assertEqual(db_ehq.who_diagnosed, ehq.who_diagnosed)
        return db_ehq

    def construct_home_dependent_questionnaire(self, dependent_living_situation='fullTimeGuardian', housemates=None,
                                               struggle_to_afford=False, participant=None, user=None):

        hq = HomeDependentQuestionnaire(dependent_living_situation=dependent_living_situation,
                                        struggle_to_afford=struggle_to_afford)

        if user is None:
            u = self.construct_user(email='homedep@questionnaire.com')
            hq.user_id = u.id
        else:
            u = user
            hq.user_id = u.id

        if participant is None:
            p = self.construct_participant(user=u, relationship=Relationship.dependent)
            hq.participant_id = p.id
        else:
            p = participant
            hq.participant_id = p.id

        db.session.add(hq)
        db.session.commit()

        if housemates is None:
            self.construct_housemate(home_dependent_questionnaire=hq)
        else:
            hq.housemates = housemates

        db_hq = db.session.query(HomeDependentQuestionnaire).filter_by(participant_id=hq.participant_id).first()
        self.assertEqual(db_hq.dependent_living_situation, hq.dependent_living_situation)
        return db_hq

    def construct_home_self_questionnaire(self, self_living_situation='alone', housemates=None,
                                          struggle_to_afford=False,
                                          participant=None, user=None):

        hq = HomeSelfQuestionnaire(self_living_situation=self_living_situation, struggle_to_afford=struggle_to_afford)

        if user is None:
            u = self.construct_user(email='homeself@questionnaire.com')
            hq.user_id = u.id
        else:
            u = user
            hq.user_id = u.id

        if participant is None:
            p = self.construct_participant(user=u, relationship=Relationship.self_participant)
            hq.participant_id = p.id
        else:
            p = participant
            hq.participant_id = p.id

        db.session.add(hq)
        db.session.commit()

        if housemates is None:
            self.construct_housemate(home_self_questionnaire=hq)
        else:
            hq.housemates = housemates

        db_hq = db.session.query(HomeSelfQuestionnaire).filter_by(participant_id=hq.participant_id).first()
        self.assertEqual(db_hq.self_living_situation, hq.self_living_situation)
        return db_hq

    def construct_housemate(self, name="Fred Fredly", relationship='bioSibling', age=23, has_autism=True,
                            home_dependent_questionnaire=None, home_self_questionnaire=None):

        h = Housemate(name=name, relationship=relationship, age=age, has_autism=has_autism)
        if home_dependent_questionnaire is not None:
            h.home_dependent_questionnaire_id = home_dependent_questionnaire.id
        if home_self_questionnaire is not None:
            h.home_self_questionnaire_id = home_self_questionnaire.id

        db.session.add(h)
        db.session.commit()

        db_h = db.session.query(Housemate).filter_by(last_updated=h.last_updated).first()
        self.assertEqual(db_h.relationship, h.relationship)
        return db_h

    def construct_identification_questionnaire(self, relationship_to_participant='adoptFather', first_name='Karl',
                                               is_first_name_preferred=False, nickname='Big K', birth_state='VA',
                                               is_english_primary=False, participant=None, user=None):

        iq = IdentificationQuestionnaire(relationship_to_participant=relationship_to_participant, first_name=first_name,
                                         is_first_name_preferred=is_first_name_preferred, nickname=nickname,
                                         birth_state=birth_state, is_english_primary=is_english_primary)
        if user is None:
            u = self.construct_user(email='ident@questionnaire.com')
            iq.user_id = u.id
        else:
            u = user
            iq.user_id = u.id

        if participant is None:
            iq.participant_id = self.construct_participant(user=u, relationship=Relationship.dependent).id
        else:
            iq.participant_id = participant.id

        db.session.add(iq)
        db.session.commit()

        db_iq = db.session.query(IdentificationQuestionnaire).filter_by(participant_id=iq.participant_id).first()
        self.assertEqual(db_iq.nickname, iq.nickname)
        return db_iq

    def construct_professional_questionnaire(self, purpose="profResearch",
                                             professional_identity=["artTher", "profOther"],
                                             professional_identity_other="Astronaut",
                                             learning_interests=["insuranceCov", "learnOther"],
                                             learning_interests_other="Data plotting using dried fruit",
                                             currently_work_with_autistic=True, previous_work_with_autistic=False,
                                             length_work_with_autistic='3 minutes', participant=None, user=None):

        pq = ProfessionalProfileQuestionnaire(purpose=purpose, professional_identity=professional_identity,
                                              professional_identity_other=professional_identity_other,
                                              learning_interests=learning_interests,
                                              learning_interests_other=learning_interests_other,
                                              currently_work_with_autistic=currently_work_with_autistic,
                                              previous_work_with_autistic=previous_work_with_autistic,
                                              length_work_with_autistic=length_work_with_autistic)
        if user is None:
            u = self.construct_user(email='prof@questionnaire.com')
            pq.user_id = u.id
        else:
            u = user
            pq.user_id = u.id

        if participant is None:
            pq.participant_id = self.construct_participant(user=u, relationship=Relationship.dependent).id
        else:
            pq.participant_id = participant.id

        db.session.add(pq)
        db.session.commit()

        db_pq = db.session.query(ProfessionalProfileQuestionnaire).filter_by(participant_id=pq.participant_id).first()
        self.assertEqual(db_pq.learning_interests, pq.learning_interests)
        return db_pq

    def construct_medication(self, symptom='symptomInsomnia', name='Magic Potion', notes='I feel better than ever!', supports_questionnaire=None):

        m = Medication(symptom=symptom, name=name, notes=notes)
        if supports_questionnaire is not None:
            m.supports_questionnaire_id = supports_questionnaire.id

        db.session.add(m)
        db.session.commit()

        db_m = db.session.query(Medication).filter_by(last_updated=m.last_updated).first()
        self.assertEqual(db_m.notes, m.notes)
        return db_m

    def construct_therapy(self, type='behavioral', timeframe='current', notes='Small steps',
                          supports_questionnaire=None):

        t = Therapy(type=type, timeframe=timeframe, notes=notes)
        if supports_questionnaire is not None:
            t.supports_questionnaire_id = supports_questionnaire.id

        db.session.add(t)
        db.session.commit()

        db_t = db.session.query(Therapy).filter_by(last_updated=t.last_updated).first()
        self.assertEqual(db_t.notes, t.notes)
        return db_t

    def construct_supports_questionnaire(self, medications=None, therapies=None, assistive_devices=None,
                                         alternative_augmentative=None, participant=None, user=None):

        sq = SupportsQuestionnaire()
        if user is None:
            u = self.construct_user(email='support@questionnaire.com')
            sq.user_id = u.id
        else:
            u = user
            sq.user_id = u.id

        if participant is None:
            sq.participant_id = self.construct_participant(user=u, relationship=Relationship.dependent).id
        else:
            sq.participant_id = participant.id

        db.session.add(sq)
        db.session.commit()

        if assistive_devices is None:
            self.construct_assistive_device(supports_questionnaire=sq)
        else:
            sq.assistive_devices = assistive_devices

        if alternative_augmentative is None:
            self.construct_alternative_augmentative(supports_questionnaire=sq)
        else:
            sq.alternative_augmentative = alternative_augmentative

        if medications is None:
            self.construct_medication(supports_questionnaire=sq)
        else:
            sq.medications = medications

        if therapies is None:
            self.construct_therapy(supports_questionnaire=sq)
        else:
            sq.therapies = therapies

        db_sq = db.session.query(SupportsQuestionnaire).filter_by(participant_id=sq.participant_id).first()
        self.assertEqual(db_sq.last_updated, sq.last_updated)
        return db_sq

    def construct_all_questionnaires(self):
        user = self.construct_user()
        participant = self.construct_participant(user=user, relationship=Relationship.dependent)
        self.construct_clinical_diagnoses_questionnaire(user=user, participant=participant)
        self.construct_contact_questionnaire(user=user, participant=participant)
        self.construct_current_behaviors_dependent_questionnaire(user=user, participant=participant)
        self.construct_current_behaviors_self_questionnaire(user=user, participant=participant)
        self.construct_demographics_questionnaire(user=user, participant=participant)
        self.construct_developmental_questionnaire(user=user, participant=participant)
        self.construct_education_dependent_questionnaire(user=user, participant=participant)
        self.construct_education_self_questionnaire(user=user, participant=participant)
        self.construct_employment_questionnaire(user=user, participant=participant)
        self.construct_evaluation_history_dependent_questionnaire(user=user, participant=participant)
        self.construct_evaluation_history_self_questionnaire(user=user, participant=participant)
        self.construct_home_dependent_questionnaire(user=user, participant=participant)
        self.construct_home_self_questionnaire(user=user, participant=participant)
        self.construct_identification_questionnaire(user=user, participant=participant)
        self.construct_professional_questionnaire(user=user, participant=participant)
        self.construct_supports_questionnaire(user=user, participant=participant)

    def test_event_basics(self):
        self.construct_event()
        r = db.session.query(Event).first()
        self.assertIsNotNone(r)
        r_id = r.id
        rv = self.app.get('/api/event/%i' % r_id,
                          follow_redirects=True,
                          content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], r_id)
        self.assertEqual(response["title"], 'A+ Event')
        self.assertEqual(response["description"], 'A delightful event destined to create rejoicing')

    def test_modify_event_basics(self):
        self.construct_event()
        r = db.session.query(Event).first()
        self.assertIsNotNone(r)
        r_id = r.id
        rv = self.app.get('/api/event/%i' % r_id, content_type="application/json")
        response = json.loads(rv.get_data(as_text=True))
        response['title'] = 'Edwarardos Lemonade and Oil Change'
        response['description'] = 'Better fluids for you and your car.'
        response['website'] = 'http://sartography.com'
        orig_date = response['last_updated']
        rv = self.app.put('/api/event/%i' % r_id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True)
        self.assertSuccess(rv)
        rv = self.app.get('/api/event/%i' % r_id, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['title'], 'Edwarardos Lemonade and Oil Change')
        self.assertEqual(response['description'], 'Better fluids for you and your car.')
        self.assertEqual(response['website'], 'http://sartography.com')
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_event(self):
        r = self.construct_event()
        r_id = r.id
        rv = self.app.get('api/event/%i' % r_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.delete('api/event/%i' % r_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.get('api/event/%i' % r_id, content_type="application/json")
        self.assertEqual(404, rv.status_code)

    def test_create_event(self):
        event = {'title': "event of events", 'description': "You need this event in your life.", 'time': "4PM sharp",
                 'ticket_cost': "$500 suggested donation"}
        rv = self.app.post('api/event', data=json.dumps(event), content_type="application/json",
                           follow_redirects=True)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['title'], 'event of events')
        self.assertEqual(response['description'], 'You need this event in your life.')
        self.assertEqual(response['time'], '4PM sharp')
        self.assertEqual(response['ticket_cost'], '$500 suggested donation')
        self.assertIsNotNone(response['id'])

    def test_location_basics(self):
        self.construct_location()
        r = db.session.query(Location).first()
        self.assertIsNotNone(r)
        r_id = r.id
        rv = self.app.get('/api/location/%i' % r_id,
                          follow_redirects=True,
                          content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], r_id)
        self.assertEqual(response["title"], 'A+ location')
        self.assertEqual(response["description"], 'A delightful location destined to create rejoicing')

    def test_modify_location_basics(self):
        self.construct_location()
        r = db.session.query(Location).first()
        self.assertIsNotNone(r)
        r_id = r.id
        rv = self.app.get('/api/location/%i' % r_id, content_type="application/json")
        response = json.loads(rv.get_data(as_text=True))
        response['title'] = 'Edwarardos Lemonade and Oil Change'
        response['description'] = 'Better fluids for you and your car.'
        response['website'] = 'http://sartography.com'
        orig_date = response['last_updated']
        rv = self.app.put('/api/location/%i' % r_id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True)
        self.assertSuccess(rv)
        rv = self.app.get('/api/location/%i' % r_id, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['title'], 'Edwarardos Lemonade and Oil Change')
        self.assertEqual(response['description'], 'Better fluids for you and your car.')
        self.assertEqual(response['website'], 'http://sartography.com')
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_location(self):
        r = self.construct_location()
        r_id = r.id
        rv = self.app.get('api/location/%i' % r_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.delete('api/location/%i' % r_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.get('api/location/%i' % r_id, content_type="application/json")
        self.assertEqual(404, rv.status_code)

    def test_create_location(self):
        location = {'title': "location of locations", 'description': "You need this location in your life."}
        rv = self.app.post('api/location', data=json.dumps(location), content_type="application/json",
                           follow_redirects=True)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['title'], 'location of locations')
        self.assertEqual(response['description'], 'You need this location in your life.')
        self.assertIsNotNone(response['id'])

    def test_resource_basics(self):
        self.construct_resource()
        r = db.session.query(StarResource).first()
        self.assertIsNotNone(r)
        r_id = r.id
        rv = self.app.get('/api/resource/%i' % r_id,
                          follow_redirects=True,
                          content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], r_id)
        self.assertEqual(response["title"], 'A+ Resource')
        self.assertEqual(response["description"], 'A delightful Resource destined to create rejoicing')

    def test_modify_resource_basics(self):
        self.construct_resource()
        r = db.session.query(StarResource).first()
        self.assertIsNotNone(r)
        r_id = r.id
        rv = self.app.get('/api/resource/%i' % r_id, content_type="application/json")
        response = json.loads(rv.get_data(as_text=True))
        response['title'] = 'Edwarardos Lemonade and Oil Change'
        response['description'] = 'Better fluids for you and your car.'
        response['website'] = 'http://sartography.com'
        orig_date = response['last_updated']
        rv = self.app.put('/api/resource/%i' % r_id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True)
        self.assertSuccess(rv)
        rv = self.app.get('/api/resource/%i' % r_id, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['title'], 'Edwarardos Lemonade and Oil Change')
        self.assertEqual(response['description'], 'Better fluids for you and your car.')
        self.assertEqual(response['website'], 'http://sartography.com')
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_resource(self):
        r = self.construct_resource()
        r_id = r.id
        rv = self.app.get('api/resource/%i' % r_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.delete('api/resource/%i' % r_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.get('api/resource/%i' % r_id, content_type="application/json")
        self.assertEqual(404, rv.status_code)

    def test_create_resource(self):
        resource = {'title': "Resource of Resources", 'description': "You need this resource in your life."}
        rv = self.app.post('api/resource', data=json.dumps(resource), content_type="application/json",
                           follow_redirects=True)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['title'], 'Resource of Resources')
        self.assertEqual(response['description'], 'You need this resource in your life.')
        self.assertIsNotNone(response['id'])

    def test_study_basics(self):
        self.construct_study()
        s = db.session.query(Study).first()
        self.assertIsNotNone(s)
        s_id = s.id
        rv = self.app.get('/api/study/%i' % s_id,
                          follow_redirects=True,
                          content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], s_id)
        self.assertEqual(response["title"], 'Fantastic Study')
        self.assertEqual(response["description"], 'A study that will go down in history')

    def test_modify_study_basics(self):
        self.construct_study()
        s = db.session.query(Study).first()
        self.assertIsNotNone(s)
        s_id = s.id
        rv = self.app.get('/api/study/%i' % s_id, content_type="application/json")
        response = json.loads(rv.get_data(as_text=True))
        response['title'] = 'Edwarardos Lemonade and Oil Change'
        response['description'] = 'Better fluids for you and your car.'
        response['benefit_description'] = 'Better fluids for you and your car, Duh.'
        orig_date = response['last_updated']
        rv = self.app.put('/api/study/%i' % s_id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True)
        self.assertSuccess(rv)
        rv = self.app.get('/api/study/%i' % s_id, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['title'], 'Edwarardos Lemonade and Oil Change')
        self.assertEqual(response['description'], 'Better fluids for you and your car.')
        self.assertEqual(response['benefit_description'], 'Better fluids for you and your car, Duh.')
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_study(self):
        s = self.construct_study()
        s_id = s.id
        rv = self.app.get('api/study/%i' % s_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.delete('api/study/%i' % s_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.get('api/study/%i' % s_id, content_type="application/json")
        self.assertEqual(404, rv.status_code)

    def test_create_study(self):
        study = {'title': "Study of Studies", 'benefit_description': "This study will change your life."}
        rv = self.app.post('api/study', data=json.dumps(study), content_type="application/json",
                           follow_redirects=True)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['title'], 'Study of Studies')
        self.assertEqual(response['benefit_description'], 'This study will change your life.')
        self.assertIsNotNone(response['id'])

    def test_organization_basics(self):
        self.construct_organization()
        o = db.session.query(Organization).first()
        self.assertIsNotNone(o)
        o_id = o.id
        rv = self.app.get('/api/organization/%i' % o_id,
                          follow_redirects=True,
                          content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], o_id)
        self.assertEqual(response["name"], 'Staunton Makerspace')
        self.assertEqual(response["description"],
                         'A place full of surprise, delight, and amazing people. And tools. Lots of exciting tools.')

    def test_modify_organization_basics(self):
        self.construct_organization()
        o = db.session.query(Organization).first()
        self.assertIsNotNone(o)
        o_id = o.id
        rv = self.app.get('/api/organization/%i' % o_id, content_type="application/json")
        response = json.loads(rv.get_data(as_text=True))
        response['name'] = 'Edwarardos Lemonade and Oil Change'
        response['description'] = 'Better fluids for you and your car.'
        orig_date = response['last_updated']
        rv = self.app.put('/api/organization/%i' % o_id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True)
        self.assertSuccess(rv)
        rv = self.app.get('/api/organization/%i' % o_id, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['name'], 'Edwarardos Lemonade and Oil Change')
        self.assertEqual(response['description'], 'Better fluids for you and your car.')
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_organization(self):
        o = self.construct_organization()
        o_id = o.id
        rv = self.app.get('api/organization/%i' % o_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.delete('api/organization/%i' % o_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.get('api/organization/%i' % o_id, content_type="application/json")
        self.assertEqual(404, rv.status_code)

    def test_create_organization(self):
        organization = {'name': "Organization of Champions", 'description': "All the best people, all the time."}
        rv = self.app.post('api/organization', data=json.dumps(organization), content_type="application/json",
                           follow_redirects=True)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['name'], 'Organization of Champions')
        self.assertEqual(response['description'], 'All the best people, all the time.')
        self.assertIsNotNone(response['id'])

    def test_category_basics(self):
        self.construct_category()
        c = db.session.query(Category).first()
        self.assertIsNotNone(c)
        c_id = c.id
        c.parent = self.construct_category(name="3d Printers")
        rv = self.app.get('/api/category/%i' % c_id,
                          follow_redirects=True,
                          content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], c_id)
        self.assertEqual(response["name"], 'Ultimakers')
        self.assertEqual(response["parent"]["name"], '3d Printers')

    def test_modify_category_basics(self):
        self.construct_category()
        c = db.session.query(Category).first()
        self.assertIsNotNone(c)
        c_id = c.id
        c.parent = self.construct_category(name="3d Printers")
        rv = self.app.get('/api/category/%i' % c_id, content_type="application/json")
        response = json.loads(rv.get_data(as_text=True))
        response['name'] = 'JellyBoxes'
        new_parent = self.construct_category(name="Strange Kitchen Gadgets")
        response['parent_id'] = new_parent.id
        rv = self.app.put('/api/category/%i' % c_id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True)
        self.assertSuccess(rv)
        rv = self.app.get('/api/category/%i' % c_id, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['name'], 'JellyBoxes')
        self.assertEqual(response['parent']['name'], 'Strange Kitchen Gadgets')

    def test_delete_category(self):
        c = self.construct_category()
        c_id = c.id
        rv = self.app.get('api/category/%i' % c_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.delete('api/category/%i' % c_id, content_type="application/json")
        self.assertSuccess(rv)

        rv = self.app.get('api/category/%i' % c_id, content_type="application/json")
        self.assertEqual(404, rv.status_code)

    def test_create_category(self):
        category = {'name': "My Favorite Things"}
        rv = self.app.post('api/category', data=json.dumps(category), content_type="application/json",
                           follow_redirects=True)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['name'], 'My Favorite Things')
        self.assertIsNotNone(response['id'])

    def test_category_has_links(self):
        self.construct_category()
        rv = self.app.get(
            '/api/category/1',
            follow_redirects=True,
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["_links"]["self"], '/api/category/1')
        self.assertEqual(response["_links"]["collection"], '/api/category')

    def test_category_has_children(self):
        c1 = self.construct_category()
        c2 = self.construct_category(name="I'm the kid", parent=c1)
        rv = self.app.get(
            '/api/category/1',
            follow_redirects=True,
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["children"][0]['id'], 2)
        self.assertEqual(response["children"][0]['name'], "I'm the kid")

    def test_category_has_parents_and_that_parent_has_no_children(self):
        c1 = self.construct_category()
        c2 = self.construct_category(name="I'm the kid", parent=c1)
        c3 = self.construct_category(name="I'm the grand kid", parent=c2)
        rv = self.app.get(
            '/api/category/3',
            follow_redirects=True,
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["parent"]['id'], 2)
        self.assertNotIn("children", response["parent"])

    def test_category_depth_is_limited(self):
        c1 = self.construct_category()
        c2 = self.construct_category(
            name="I'm the kid", parent=c1)
        c3 = self.construct_category(
            name="I'm the grand kid",
            parent=c2)
        c4 = self.construct_category(
            name="I'm the great grand kid",
            parent=c3)

        rv = self.app.get(
            '/api/category',
            follow_redirects=True,
            content_type="application/json")

        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))

        self.assertEqual(1, len(response))
        self.assertEqual(1, len(response[0]["children"]))

    def test_get_event_by_category(self):
        c = self.construct_category()
        ev = self.construct_event()
        rc = ResourceCategory(resource_id=ev.id, category=c, type='event')
        db.session.add(rc)
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i/event' % c.id,
            content_type="application/json",
            headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))
        self.assertEqual(ev.id, response[0]["id"])
        self.assertEqual(ev.description, response[0]["resource"]["description"])

    def test_get_event_by_category_includes_category_details(self):
        c = self.construct_category(name="c1")
        c2 = self.construct_category(name="c2")
        ev = self.construct_event()
        rc = ResourceCategory(resource_id=ev.id, category=c, type='event')
        rc2 = ResourceCategory(resource_id=ev.id, category=c2, type='event')
        db.session.add_all([rc, rc2])
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i/event' % c.id,
            content_type="application/json",
            headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(ev.id, response[0]["id"])
        self.assertEqual(2,
                         len(response[0]["resource"]["resource_categories"]))
        self.assertEqual(
            "c1", response[0]["resource"]["resource_categories"][0]["category"]
            ["name"])

    def test_category_event_count(self):
        c = self.construct_category()
        ev = self.construct_event()
        rec = self.construct_resource()
        rc = ResourceCategory(resource_id=ev.id, category=c, type='event')
        rc2 = ResourceCategory(resource_id=rec.id, category=c, type='resource')
        db.session.add_all([rc, rc2])
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i' % c.id, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, response["event_count"])

    def test_get_category_by_event(self):
        c = self.construct_category()
        ev = self.construct_event()
        rc = ResourceCategory(resource_id=ev.id, category=c, type='event')
        db.session.add(rc)
        db.session.commit()
        rv = self.app.get(
            '/api/event/%i/category' % ev.id,
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))
        self.assertEqual(c.id, response[0]["id"])
        self.assertEqual(c.name, response[0]["category"]["name"])

    def test_add_category_to_event(self):
        c = self.construct_category()
        ev = self.construct_event()

        ec_data = {"resource_id": ev.id, "category_id": c.id}

        rv = self.app.post(
            '/api/resource_category',
            data=json.dumps(ec_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(c.id, response["category_id"])
        self.assertEqual(ev.id, response["resource_id"])

    def test_set_all_categories_on_event(self):
        c1 = self.construct_category(name="c1")
        c2 = self.construct_category(name="c2")
        c3 = self.construct_category(name="c3")
        ev = self.construct_event()

        ec_data = [
            {
                "category_id": c1.id
            },
            {
                "category_id": c2.id
            },
            {
                "category_id": c3.id
            },
        ]
        rv = self.app.post(
            '/api/event/%i/category' % ev.id,
            data=json.dumps(ec_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(3, len(response))

        ec_data = [{"category_id": c1.id}]
        rv = self.app.post(
            '/api/event/%i/category' % ev.id,
            data=json.dumps(ec_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))

    def test_remove_category_from_event(self):
        self.test_add_category_to_event()
        rv = self.app.delete('/api/resource_category/%i' % 1)
        self.assertSuccess(rv)
        rv = self.app.get(
            '/api/event/%i/category' % 1, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(0, len(response))

    def test_get_location_by_category(self):
        c = self.construct_category()
        loc = self.construct_location()
        rc = ResourceCategory(resource_id=loc.id, category=c, type='location')
        db.session.add(rc)
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i/location' % c.id,
            content_type="application/json",
            headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))
        self.assertEqual(loc.id, response[0]["id"])
        self.assertEqual(loc.description, response[0]["resource"]["description"])

    def test_get_location_by_category_includes_category_details(self):
        c = self.construct_category(name="c1")
        c2 = self.construct_category(name="c2")
        loc = self.construct_location()
        rc = ResourceCategory(resource_id=loc.id, category=c, type='location')
        rc2 = ResourceCategory(resource_id=loc.id, category=c2, type='location')
        db.session.add_all([rc, rc2])
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i/location' % c.id,
            content_type="application/json",
            headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(loc.id, response[0]["id"])
        self.assertEqual(2,
                         len(response[0]["resource"]["resource_categories"]))
        self.assertEqual(
            "c1", response[0]["resource"]["resource_categories"][0]["category"]
            ["name"])

    def test_category_location_count(self):
        c = self.construct_category()
        loc = self.construct_location()
        rc = ResourceCategory(resource_id=loc.id, category=c, type='location')
        db.session.add(rc)
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i' % c.id, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, response["location_count"])

    def test_get_category_by_location(self):
        c = self.construct_category()
        loc = self.construct_location()
        rc = ResourceCategory(resource_id=loc.id, category=c, type='location')
        db.session.add(rc)
        db.session.commit()
        rv = self.app.get(
            '/api/location/%i/category' % loc.id,
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))
        self.assertEqual(c.id, response[0]["id"])
        self.assertEqual(c.name, response[0]["category"]["name"])

    def test_add_category_to_location(self):
        c = self.construct_category()
        loc = self.construct_location()

        rc_data = {"resource_id": loc.id, "category_id": c.id}

        rv = self.app.post(
            '/api/resource_category',
            data=json.dumps(rc_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(c.id, response["category_id"])
        self.assertEqual(loc.id, response["resource_id"])

    def test_set_all_categories_on_location(self):
        c1 = self.construct_category(name="c1")
        c2 = self.construct_category(name="c2")
        c3 = self.construct_category(name="c3")
        loc = self.construct_location()

        lc_data = [
            {
                "category_id": c1.id
            },
            {
                "category_id": c2.id
            },
            {
                "category_id": c3.id
            },
        ]
        rv = self.app.post(
            '/api/location/%i/category' % loc.id,
            data=json.dumps(lc_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(3, len(response))

        lc_data = [{"category_id": c1.id}]
        rv = self.app.post(
            '/api/location/%i/category' % loc.id,
            data=json.dumps(lc_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))

    def test_remove_category_from_location(self):
        self.test_add_category_to_location()
        rv = self.app.delete('/api/resource_category/%i' % 1)
        self.assertSuccess(rv)
        rv = self.app.get(
            '/api/location/%i/category' % 1, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(0, len(response))

    def test_get_resource_by_category(self):
        c = self.construct_category()
        r = self.construct_resource()
        cr = ResourceCategory(resource=r, category=c, type='resource')
        db.session.add(cr)
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i/resource' % c.id,
            content_type="application/json",
            headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))
        self.assertEqual(r.id, response[0]["id"])
        self.assertEqual(r.description, response[0]["resource"]["description"])

    def test_get_resource_by_category_includes_category_details(self):
        c = self.construct_category(name="c1")
        c2 = self.construct_category(name="c2")
        r = self.construct_resource()
        cr = ResourceCategory(resource=r, category=c, type='resource')
        cr2 = ResourceCategory(resource=r, category=c2, type='resource')
        db.session.add_all([cr, cr2])
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i/resource' % c.id,
            content_type="application/json",
            headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(r.id, response[0]["id"])
        self.assertEqual(2,
                         len(response[0]["resource"]["resource_categories"]))
        self.assertEqual(
            "c1", response[0]["resource"]["resource_categories"][0]["category"]
            ["name"])

    def test_category_resource_count(self):
        c = self.construct_category()
        r = self.construct_resource()
        cr = ResourceCategory(resource=r, category=c, type='resource')
        db.session.add(cr)
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i' % c.id, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, response["resource_count"])

    def test_get_category_by_resource(self):
        c = self.construct_category()
        r = self.construct_resource()
        cr = ResourceCategory(resource=r, category=c, type='resource')
        db.session.add(cr)
        db.session.commit()
        rv = self.app.get(
            '/api/resource/%i/category' % r.id,
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))
        self.assertEqual(c.id, response[0]["id"])
        self.assertEqual(c.name, response[0]["category"]["name"])

    def test_add_category_to_resource(self):
        c = self.construct_category()
        r = self.construct_resource()

        rc_data = {"resource_id": r.id, "category_id": c.id}

        rv = self.app.post(
            '/api/resource_category',
            data=json.dumps(rc_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(c.id, response["category_id"])
        self.assertEqual(r.id, response["resource_id"])

    def test_set_all_categories_on_resource(self):
        c1 = self.construct_category(name="c1")
        c2 = self.construct_category(name="c2")
        c3 = self.construct_category(name="c3")
        r = self.construct_resource()

        rc_data = [
            {
                "category_id": c1.id
            },
            {
                "category_id": c2.id
            },
            {
                "category_id": c3.id
            },
        ]
        rv = self.app.post(
            '/api/resource/%i/category' % r.id,
            data=json.dumps(rc_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(3, len(response))

        rc_data = [{"category_id": c1.id}]
        rv = self.app.post(
            '/api/resource/%i/category' % r.id,
            data=json.dumps(rc_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))

    def test_remove_category_from_resource(self):
        self.test_add_category_to_resource()
        rv = self.app.delete('/api/resource_category/%i' % 1)
        self.assertSuccess(rv)
        rv = self.app.get(
            '/api/resource/%i/category' % 1, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(0, len(response))

    def test_get_study_by_category(self):
        c = self.construct_category()
        s = self.construct_study()
        cs = StudyCategory(study=s, category=c)
        db.session.add(cs)
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i/study' % c.id,
            content_type="application/json",
            headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))
        self.assertEqual(s.id, response[0]["id"])
        self.assertEqual(s.description, response[0]["study"]["description"])

    def test_get_study_by_category_includes_category_details(self):
        c = self.construct_category(name="c1")
        c2 = self.construct_category(name="c2")
        s = self.construct_study()
        cs = StudyCategory(study=s, category=c)
        cs2 = StudyCategory(study=s, category=c2)
        db.session.add_all([cs, cs2])
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i/study' % c.id,
            content_type="application/json",
            headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(s.id, response[0]["id"])
        self.assertEqual(2,
                         len(response[0]["study"]["study_categories"]))
        self.assertEqual(
            "c1", response[0]["study"]["study_categories"][0]["category"]
            ["name"])

    def test_category_study_count(self):
        c = self.construct_category()
        s = self.construct_study()
        cs = StudyCategory(study=s, category=c)
        db.session.add(cs)
        db.session.commit()
        rv = self.app.get(
            '/api/category/%i' % c.id, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, response["study_count"])

    def test_get_category_by_study(self):
        c = self.construct_category()
        s = self.construct_study()
        cs = StudyCategory(study=s, category=c)
        db.session.add(cs)
        db.session.commit()
        rv = self.app.get(
            '/api/study/%i/category' % s.id,
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))
        self.assertEqual(c.id, response[0]["id"])
        self.assertEqual(c.name, response[0]["category"]["name"])

    def test_add_category_to_study(self):
        c = self.construct_category()
        s = self.construct_study()

        sc_data = {"study_id": s.id, "category_id": c.id}

        rv = self.app.post(
            '/api/study_category',
            data=json.dumps(sc_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(c.id, response["category_id"])
        self.assertEqual(s.id, response["study_id"])

    def test_set_all_categories_on_study(self):
        c1 = self.construct_category(name="c1")
        c2 = self.construct_category(name="c2")
        c3 = self.construct_category(name="c3")
        s = self.construct_study()

        sc_data = [
            {
                "category_id": c1.id
            },
            {
                "category_id": c2.id
            },
            {
                "category_id": c3.id
            },
        ]
        rv = self.app.post(
            '/api/study/%i/category' % s.id,
            data=json.dumps(sc_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(3, len(response))

        sc_data = [{"category_id": c1.id}]
        rv = self.app.post(
            '/api/study/%i/category' % s.id,
            data=json.dumps(sc_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))

    def test_remove_category_from_study(self):
        self.test_add_category_to_study()
        rv = self.app.delete('/api/study_category/%i' % 1)
        self.assertSuccess(rv)
        rv = self.app.get(
            '/api/study/%i/category' % 1, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(0, len(response))

    def test_add_investigator_to_study(self):
        i = self.construct_investigator()
        s = self.construct_study()

        si_data = {"study_id": s.id, "investigator_id": i.id}

        rv = self.app.post(
            '/api/study_investigator',
            data=json.dumps(si_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(i.id, response["investigator_id"])
        self.assertEqual(s.id, response["study_id"])

    def test_set_all_investigators_on_study(self):
        i1 = self.construct_investigator(name="person1")
        i2 = self.construct_investigator(name="person2")
        i3 = self.construct_investigator(name="person3")
        s = self.construct_study()

        si_data = [
            {"investigator_id": i1.id},
            {"investigator_id": i2.id},
            {"investigator_id": i3.id},
        ]
        rv = self.app.post(
            '/api/study/%i/investigator' % s.id,
            data=json.dumps(si_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(3, len(response))

        si_data = [{"investigator_id": i1.id}]
        rv = self.app.post(
            '/api/study/%i/investigator' % s.id,
            data=json.dumps(si_data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, len(response))

    def test_remove_investigator_from_study(self):
        self.test_add_investigator_to_study()
        rv = self.app.delete('/api/study_investigator/%i' % 1)
        self.assertSuccess(rv)
        rv = self.app.get(
            '/api/study/%i/investigator' % 1, content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(0, len(response))

    def test_investigator_basics(self):
        self.construct_investigator()
        i = db.session.query(Investigator).first()
        self.assertIsNotNone(i)
        i_id = i.id
        rv = self.app.get('/api/investigator/%i' % i_id,
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], i_id)
        self.assertEqual(response["name"], i.name)

    def test_modify_investigator_basics(self):
        self.construct_investigator()
        i = db.session.query(Investigator).first()
        self.assertIsNotNone(i)

        rv = self.app.get('/api/investigator/%i' % i.id, content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        response['title'] = 'dungeon master'
        orig_date = response['last_updated']
        rv = self.app.put('/api/investigator/%i' % i.id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True, headers=self.logged_in_headers())
        self.assertSuccess(rv)

        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['title'], 'dungeon master')
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_investigator(self):
        i = self.construct_investigator()
        i_id = i.id

        rv = self.app.get('api/investigator/%i' % i_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/investigator/%i' % i_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/investigator/%i' % i_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_investigator(self):
        investigator = {'name': "Tara Tarantula", 'title': "Assistant Professor of Arachnology"}
        rv = self.app.post('api/investigator', data=json.dumps(investigator), content_type="application/json",
                           headers=self.logged_in_headers(), follow_redirects=True)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['name'], 'Tara Tarantula')
        self.assertEqual(response['title'], 'Assistant Professor of Arachnology')
        self.assertIsNotNone(response['id'])

    def search(self, query, user=None):
        """Executes a query as the given user, returning the resulting search results object."""
        rv = self.app.post(
            '/api/search',
            data=json.dumps(query),
            follow_redirects=True,
            content_type="application/json",
            headers=self.logged_in_headers(user))
        self.assertSuccess(rv)
        return json.loads(rv.get_data(as_text=True))

    def search_anonymous(self, query):
        """Executes a query as an anonymous user, returning the resulting search results object."""
        rv = self.app.post(
            '/api/search',
            data=json.dumps(query),
            follow_redirects=True,
            content_type="application/json")
        self.assertSuccess(rv)
        return json.loads(rv.get_data(as_text=True))

    def test_search_basics(self):
        elastic_index.clear()
        rainbow_query = {'words': 'rainbows', 'filters': []}
        world_query = {'words': 'world', 'filters': []}
        search_results = self.search(rainbow_query)
        self.assertEqual(0, len(search_results["hits"]))
        search_results = self.search(world_query)
        self.assertEqual(0, len(search_results["hits"]))

        # test that elastic resource is created with post
        resource = {'title': "space unicorn", 'description': "delivering rainbows"}
        rv = self.app.post('api/resource', data=json.dumps(resource), content_type="application/json",
                           follow_redirects=True)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))

        search_results = self.search(rainbow_query)
        self.assertEqual(1, len(search_results["hits"]))
        self.assertEqual(search_results['hits'][0]['id'], response['id'])
        self.assertEqual(search_results['hits'][0]['title'], response['title'])
        self.assertEqual(search_results['hits'][0]['type'], "resource")
        self.assertEqual(search_results['hits'][0]['highlights'], "delivering <em>rainbows</em>")
        self.assertIsNotNone(search_results['hits'][0]['last_updated'])
        search_results = self.search(world_query)
        self.assertEqual(0, len(search_results["hits"]))

    def test_study_search_basics(self):
        elastic_index.clear()
        rainbow_query = {'words': 'umbrellas', 'filters': []}
        world_query = {'words': 'universe', 'filters': []}
        search_results = self.search(rainbow_query)
        self.assertEqual(0, len(search_results["hits"]))
        search_results = self.search(world_query)
        self.assertEqual(0, len(search_results["hits"]))

        # test that elastic resource is created with post
        study = {'title': "space platypus", 'description': "delivering umbrellas"}
        rv = self.app.post('api/study', data=json.dumps(study), content_type="application/json",
                           follow_redirects=True)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))

        search_results = self.search(rainbow_query)
        self.assertEqual(1, len(search_results["hits"]))
        self.assertEqual(search_results['hits'][0]['id'], response['id'])
        search_results = self.search(world_query)
        self.assertEqual(0, len(search_results["hits"]))

    def test_modify_resource_search_basics(self):
        elastic_index.clear()
        rainbow_query = {'words': 'rainbows', 'filters': []}
        world_query = {'words': 'world', 'filters': []}
        # create the resource
        resource = self.construct_resource(
            title='space unicorn', description="delivering rainbows")
        # test that it indeed exists
        rv = self.app.get('/api/resource/%i' % resource.id, content_type="application/json",
                          follow_redirects=True)
        self.assertSuccess(rv)

        search_results = self.search(rainbow_query)
        self.assertEqual(1, len(search_results["hits"]))
        self.assertEqual(search_results['hits'][0]['id'], resource.id)

        response = json.loads(rv.get_data(as_text=True))
        response['description'] = 'all around the world'
        rv = self.app.put('/api/resource/%i' % resource.id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True)
        self.assertSuccess(rv)

        search_results = self.search(rainbow_query)
        self.assertEqual(0, len(search_results["hits"]))
        search_results = self.search(world_query)
        self.assertEqual(1, len(search_results["hits"]))
        self.assertEqual(resource.id, search_results['hits'][0]['id'])

    def test_delete_search_item(self):
        rainbow_query = {'words': 'rainbows', 'filters': []}
        world_query = {'words': 'world', 'filters': []}
        resource = self.construct_resource(
            title='space unicorn', description="delivering rainbows")
        search_results = self.search(rainbow_query)
        self.assertEqual(1, len(search_results["hits"]))
        elastic_index.remove_document(resource, 'Resource')
        search_results = self.search(world_query)
        self.assertEqual(0, len(search_results["hits"]))

    def test_user_basics(self):
        self.construct_user()
        u = db.session.query(User).first()
        self.assertIsNotNone(u)
        u_id = u.id
        headers = self.logged_in_headers(u)
        rv = self.app.get('/api/user/%i' % u_id,
                          follow_redirects=True,
                          content_type="application/json", headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], u_id)
        self.assertEqual(response["email"], 'stan@staunton.com')

    def test_modify_user_basics(self):
        self.construct_user()
        u = db.session.query(User).first()
        admin_headers = self.logged_in_headers()
        user_headers = self.logged_in_headers(u)
        self.assertIsNotNone(u)

        # A user should be able to access and modify their user record, with the exception of making themselves Admin
        rv = self.app.get('/api/user/%i' % u.id, content_type="application/json", headers=user_headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        response['email'] = 'ed@edwardos.com'
        orig_date = response['last_updated']
        rv = self.app.put('/api/user/%i' % u.id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True, headers=user_headers)
        self.assertSuccess(rv)

        # Only Admin users can make other admin users
        response['role'] = 'admin'
        rv = self.app.put('/api/user/%i' % u.id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True, headers=admin_headers)
        self.assertSuccess(rv)

        rv = self.app.get('/api/user/%i' % u.id, content_type="application/json", headers=user_headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['email'], 'ed@edwardos.com')
        self.assertEqual(response['role'], 'admin')
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_user(self):
        u = self.construct_user()
        u_id = u.id

        rv = self.app.get('api/user/%i' % u_id, content_type="application/json")
        self.assertEqual(401, rv.status_code)
        rv = self.app.get('api/user/%i' % u_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/user/%i' % u_id, content_type="application/json", headers=None)
        self.assertEqual(401, rv.status_code)
        rv = self.app.delete('api/user/%i' % u_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/user/%i' % u_id, content_type="application/json")
        self.assertEqual(401, rv.status_code)
        rv = self.app.get('api/user/%i' % u_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_user(self):
        user = {'email': "tara@spiders.org"}
        rv = self.app.post('api/user', data=json.dumps(user), content_type="application/json",
                           headers=self.logged_in_headers(), follow_redirects=True)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['email'], 'tara@spiders.org')
        self.assertIsNotNone(response['id'])

    def test_create_user_with_bad_role(self):
        user = {'email': "tara@spiders.org", 'role': 'web_weaver'}

        # post should change unknown role to 'user'
        rv = self.app.post('/api/user', data=json.dumps(user),
                           content_type="application/json", follow_redirects=True, headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['role'], 'user')

    def test_non_admin_cannot_create_admin_user(self):
        u = self.construct_user()

        # post should make role 'user'
        new_admin_user = {'email': "tara@spiders.org", 'role': 'admin'}
        rv = self.app.post('/api/user', data=json.dumps(new_admin_user),
                           content_type="application/json", follow_redirects=True, headers=self.logged_in_headers(user=u))
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['role'], 'user')
        new_id = response['id']

        # put as non-admin user should keep role as 'user'
        rv = self.app.put('/api/user/%i' % u.id, data=json.dumps({'email': u.email, 'role': 'admin'}),
                           content_type="application/json", follow_redirects=True, headers=self.logged_in_headers(user=u))
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['role'], 'user')

        # put as admin user should allow to make role 'admin'
        rv = self.app.put('/api/user/%i' % new_id, data=json.dumps(new_admin_user),
                           content_type="application/json", follow_redirects=True, headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['role'], 'admin')

    def logged_in_headers(self, user=None):
        # If no user is provided, generate a dummy Admin user
        if not user:
            user = User(
                id=7,
                email="admin@admin.org",
                password="myPass457",
                email_verified=True,
                role=Role.admin)

        # Add user if it's not already in database
        existing_user = None
        if user.id:
            existing_user = User.query.filter_by(id=user.id).first()
        if not existing_user and user.email:
            existing_user = User.query.filter_by(email=user.email).first()

        if not existing_user:
            db.session.add(user)
            db.session.commit()

        data = {
            'email': user.email,
            'password': 'myPass457'
        }

        rv = self.app.post(
            '/api/login_password',
            data=json.dumps(data),
            content_type="application/json")

        db_user = User.query.filter_by(email=user.email).first()
        return dict(
            Authorization='Bearer ' + db_user.encode_auth_token().decode())

    def test_create_user_with_password(self, id=8, email="tyrion@got.com", role=Role.user, password="peterpass"):
        data = {
            "id": id,
            "email": email
        }
        rv = self.app.post(
            '/api/user',
            data=json.dumps(data),
            follow_redirects=True,
            headers=self.logged_in_headers(),
            content_type="application/json")
        self.assertSuccess(rv)
        user = User.query.filter_by(id=id).first()
        user.password = password
        user.role = role
        db.session.add(user)
        db.session.commit()

        rv = self.app.get(
            '/api/user/%i' % user.id,
            content_type="application/json",
            headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(email, response["email"])
        self.assertEqual(role.name, response["role"])
        self.assertEqual(True, user.is_correct_password(password))

        return user

    def test_login_user(self):
        user = self.test_create_user_with_password()
        data = {"email": "tyrion@got.com", "password": "peterpass"}
        # Login shouldn't work with email not yet verified
        rv = self.app.post(
            '/api/login_password',
            data=json.dumps(data),
            content_type="application/json")
        self.assertEqual(400, rv.status_code)

        user.email_verified = True
        rv = self.app.post(
            '/api/login_password',
            data=json.dumps(data),
            content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertIsNotNone(response["token"])

        return user

    def test_get_current_user(self):
        """ Test for the current user status """
        user = self.test_login_user()

        # Now get the user back.
        response = self.app.get(
            '/api/session',
            headers=dict(
                Authorization='Bearer ' + user.encode_auth_token().decode()))
        self.assertSuccess(response)
        return json.loads(response.data.decode())

    def decode(self, encoded_words):
        """
        Useful for checking the content of email messages
        (which we store in an array for testing)
        """
        encoded_word_regex = r'=\?{1}(.+)\?{1}([b|q])\?{1}(.+)\?{1}='
        charset, encoding, encoded_text = re.match(encoded_word_regex,
                                                   encoded_words).groups()
        if encoding is 'b':
            byte_string = base64.b64decode(encoded_text)
        elif encoding is 'q':
            byte_string = quopri.decodestring(encoded_text)
        text = byte_string.decode(charset)
        text = text.replace("_", " ")
        return text

    def test_register_sends_email(self):
        message_count = len(TEST_MESSAGES)
        self.test_create_user_with_password()
        self.assertGreater(len(TEST_MESSAGES), message_count)
        self.assertEqual("STAR Drive: Confirm Email",
                         self.decode(TEST_MESSAGES[-1]['subject']))

        logs = EmailLog.query.all()
        self.assertIsNotNone(logs[-1].tracking_code)

    def test_forgot_password_sends_email(self):
        user = self.test_create_user_with_password()
        message_count = len(TEST_MESSAGES)
        data = {"email": user.email}
        rv = self.app.post(
            '/api/forgot_password',
            data=json.dumps(data),
            content_type="application/json")
        self.assertSuccess(rv)
        self.assertGreater(len(TEST_MESSAGES), message_count)
        self.assertEqual("STAR Drive: Password Reset Email",
                         self.decode(TEST_MESSAGES[-1]['subject']))

        logs = EmailLog.query.all()
        self.assertIsNotNone(logs[-1].tracking_code)

    def test_participant_relationships(self):
        u = self.construct_user()
        participant = self.construct_participant(user=u, relationship=Relationship.self_participant)
        guardian = self.construct_participant(user=u, relationship=Relationship.self_guardian)
        dependent = self.construct_participant(user=u, relationship=Relationship.dependent)
        professional = self.construct_participant(user=u, relationship=Relationship.self_professional)
        rv = self.app.get('/api/user/%i' % u.id,
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], u.id)
        self.assertEqual(len(response["participants"]), 4)
        self.assertEqual(response["participants"][0]["id"], participant.id)
        self.assertEqual(response["participants"][0]["relationship"], 'self_participant')
        self.assertEqual(response["participants"][1]["id"], guardian.id)
        self.assertEqual(response["participants"][1]["relationship"], 'self_guardian')
        self.assertEqual(response["participants"][2]["id"], dependent.id)
        self.assertEqual(response["participants"][2]["relationship"], 'dependent')
        self.assertEqual(response["participants"][3]["id"], professional.id)
        self.assertEqual(response["participants"][3]["relationship"], 'self_professional')

    def test_participant_basics(self):
        self.construct_participant(user=self.construct_user(), relationship=Relationship.dependent)
        p = db.session.query(Participant).first()
        self.assertIsNotNone(p)
        p_id = p.id
        rv = self.app.get('/api/participant/%i' % p_id,
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], p_id)
        self.assertEqual(response["relationship"], p.relationship.name)

    def test_modify_participant_you_do_not_own(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        good_headers = self.logged_in_headers(u)

        p = db.session.query(Participant).first()
        odd_user = User(
            email="frankie@badfella.fr",
            password="h@corleet",
            email_verified=True,
            role=Role.user)
        participant = {'first_name': "Lil' Johnny", 'last_name': "Tables"}
        rv = self.app.put('/api/participant/%i' % p.id, data=json.dumps(participant), content_type="application/json",
                          follow_redirects=True)
        self.assertEqual(401, rv.status_code, "you have to be logged in to edit participant.")
        rv = self.app.put('/api/participant/%i' % p.id, data=json.dumps(participant), content_type="application/json",
                          follow_redirects=True, headers=self.logged_in_headers(odd_user))
        self.assertEqual(400, rv.status_code, "you have to have a relationship with the user to do stuff.")
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual("unrelated_participant", response['code'])
        rv = self.app.put('/api/participant/%i' % p.id, data=json.dumps(participant), content_type="application/json",
                          follow_redirects=True, headers=good_headers)
        self.assertEqual(200, rv.status_code, "The owner can edit the user.")

    def test_modify_participant_you_do_own(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_guardian)
        headers = self.logged_in_headers(u)
        participant = {'id': 567}
        rv = self.app.put('/api/participant/%i' % p.id, data=json.dumps(participant), content_type="application/json",
                          follow_redirects=True, headers=headers)
        self.assertSuccess(rv)
        p = db.session.query(Participant).filter_by(id=p.id).first()
        self.assertEqual(567, p.id)

    def test_modify_participant_basics_admin(self):
        self.construct_participant(user=self.construct_user(), relationship=Relationship.dependent)
        user2 = self.construct_user(email="theotherguy@stuff.com")
        p = db.session.query(Participant).first()
        self.assertIsNotNone(p)
        p_id = p.id
        rv = self.app.get('/api/participant/%i' % p_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        response['user_id'] = user2.id
        orig_date = response['last_updated']
        rv = self.app.put('/api/participant/%i' % p_id, data=json.dumps(response), content_type="application/json",
                          follow_redirects=True, headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/participant/%i' % p_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['user_id'], user2.id)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_participant(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.dependent)
        p_id = p.id
        rv = self.app.get('api/participant/%i' % p_id, content_type="application/json")
        self.assertEqual(401, rv.status_code)
        rv = self.app.get('api/participant/%i' % p_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/participant/%i' % p_id, content_type="application/json")
        self.assertEqual(401, rv.status_code)
        rv = self.app.delete('api/participant/%i' % p_id, content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/participant/%i' % p_id, content_type="application/json")
        self.assertEqual(401, rv.status_code)
        rv = self.app.get('api/participant/%i' % p_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_participant(self):
        p = {'id': 7, 'relationship': 'self_participant'}
        rv = self.app.post('/api/session/participant', data=json.dumps(p), content_type="application/json",
                           follow_redirects=True)
        self.assertEqual(401, rv.status_code, "you can't create a participant without an account.")

        rv = self.app.post(
            '/api/session/participant', data=json.dumps(p),
            content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        participant = db.session.query(Participant).filter_by(id=p['id']).first()

        self.assertIsNotNone(participant.id)
        self.assertIsNotNone(participant.user_id)

    def test_create_participant_to_have_bad_relationship(self):
        participant = {'id': 234, 'relationship': 'free_loader'}
        rv = self.app.post('/api/session/participant', data=json.dumps(participant),
                           content_type="application/json", follow_redirects=True, headers=self.logged_in_headers())
        self.assertEqual(400, rv.status_code, "you can't create a participant using an invalid relationship")
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["code"], "unknown_relationship")

    def test_get_participant_by_user(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        db.session.commit()
        rv = self.app.get(
            '/api/user/%i' % u.id,
            content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(u.id, response['id'])
        self.assertEqual(1, len(response['participants']))
        self.assertEqual(p.relationship.name, response['participants'][0]["relationship"])

    def test_clinical_diagnoses_questionnaire_basics(self):
        self.construct_clinical_diagnoses_questionnaire()
        cq = db.session.query(ClinicalDiagnosesQuestionnaire).first()
        self.assertIsNotNone(cq)
        cq_id = cq.id
        rv = self.app.get('/api/q/clinical_diagnoses_questionnaire/%i' % cq_id,
                          follow_redirects=True,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], cq_id)
        self.assertEqual(response["medical"], cq.medical)
        self.assertEqual(response["genetic"], cq.genetic)

    def test_modify_clinical_diagnoses_questionnaire_basics(self):
        self.construct_clinical_diagnoses_questionnaire()
        cq = db.session.query(ClinicalDiagnosesQuestionnaire).first()
        self.assertIsNotNone(cq)
        cq_id = cq.id
        rv = self.app.get('/api/q/clinical_diagnoses_questionnaire/%i' % cq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['developmental'] = ['intellectual']
        response['mental_health'] = ['depression']
        response['medical'] = ['gastrointestinal']
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/clinical_diagnoses_questionnaire/%i' % cq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/clinical_diagnoses_questionnaire/%i' % cq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['developmental'], ['intellectual'])
        self.assertEqual(response['mental_health'], ['depression'])
        self.assertEqual(response['medical'], ['gastrointestinal'])
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_clinical_diagnoses_questionnaire(self):
        cq = self.construct_clinical_diagnoses_questionnaire()
        cq_id = cq.id
        rv = self.app.get('api/q/clinical_diagnoses_questionnaire/%i' % cq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/clinical_diagnoses_questionnaire/%i' % cq_id, content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/clinical_diagnoses_questionnaire/%i' % cq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_clinical_diagnoses_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        clinical_diagnoses_questionnaire = {'medical': ['seizure'], 'genetic': ['fragileX'], 'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/clinical_diagnoses_questionnaire',
                           data=json.dumps(clinical_diagnoses_questionnaire),
                           content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['medical'], ['seizure'])
        self.assertEqual(response['genetic'], ['fragileX'])
        self.assertIsNotNone(response['id'])

    def test_contact_questionnaire_basics(self):
        self.construct_contact_questionnaire()
        cq = db.session.query(ContactQuestionnaire).first()
        self.assertIsNotNone(cq)
        cq_id = cq.id
        rv = self.app.get('/api/q/contact_questionnaire/%i' % cq_id,
                          follow_redirects=True,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], cq_id)
        self.assertEqual(response["phone"], cq.phone)
        self.assertEqual(response["marketing_channel"], cq.marketing_channel)

    def test_modify_contact_questionnaire_basics(self):
        self.construct_contact_questionnaire()
        cq = db.session.query(ContactQuestionnaire).first()
        self.assertIsNotNone(cq)
        cq_id = cq.id
        rv = self.app.get('/api/q/contact_questionnaire/%i' % cq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['phone'] = '123-456-7890'
        response['zip'] = 22345
        response['marketing_channel'] = 'flyer'
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/contact_questionnaire/%i' % cq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/contact_questionnaire/%i' % cq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['phone'], '123-456-7890')
        self.assertEqual(response['zip'], 22345)
        self.assertEqual(response['marketing_channel'], 'flyer')
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_contact_questionnaire(self):
        cq = self.construct_contact_questionnaire()
        cq_id = cq.id
        rv = self.app.get('api/q/contact_questionnaire/%i' % cq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/contact_questionnaire/%i' % cq_id, content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/contact_questionnaire/%i' % cq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_contact_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        contact_questionnaire = {'phone': "123-456-7890", 'marketing_channel': "Subway sign", 'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/contact_questionnaire', data=json.dumps(contact_questionnaire),
                           content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['phone'], '123-456-7890')
        self.assertEqual(response['marketing_channel'], 'Subway sign')
        self.assertIsNotNone(response['id'])

    def test_current_behaviors_dependent_questionnaire_basics(self):
        self.construct_current_behaviors_dependent_questionnaire()
        cbdq = db.session.query(CurrentBehaviorsDependentQuestionnaire).first()
        self.assertIsNotNone(cbdq)
        cbdq_id = cbdq.id
        rv = self.app.get('/api/q/current_behaviors_dependent_questionnaire/%i' % cbdq_id,
                          follow_redirects=True,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], cbdq_id)
        self.assertEqual(response["concerning_behaviors"], cbdq.concerning_behaviors)
        self.assertEqual(response["has_academic_difficulties"], cbdq.has_academic_difficulties)

    def test_modify_current_behaviors_dependent_questionnaire_basics(self):
        self.construct_current_behaviors_dependent_questionnaire()
        cbdq = db.session.query(CurrentBehaviorsDependentQuestionnaire).first()
        self.assertIsNotNone(cbdq)
        cbdq_id = cbdq.id
        rv = self.app.get('/api/q/current_behaviors_dependent_questionnaire/%i' % cbdq_id,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['dependent_verbal_ability'] = 'nonVerbal'
        response['concerning_behaviors'] = ['elopement']
        response['has_academic_difficulties'] = False
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/current_behaviors_dependent_questionnaire/%i' % cbdq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/current_behaviors_dependent_questionnaire/%i' % cbdq_id,
                          content_type="application/json",headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['dependent_verbal_ability'], 'nonVerbal')
        self.assertEqual(response['concerning_behaviors'], ['elopement'])
        self.assertEqual(response['has_academic_difficulties'], False)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_current_behaviors_dependent_questionnaire(self):
        cbdq = self.construct_current_behaviors_dependent_questionnaire()
        cbdq_id = cbdq.id
        rv = self.app.get('api/q/current_behaviors_dependent_questionnaire/%i' % cbdq_id,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/current_behaviors_dependent_questionnaire/%i' % cbdq_id,
                             content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/current_behaviors_dependent_questionnaire/%i' % cbdq_id,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_current_behaviors_dependent_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.dependent)
        headers = self.logged_in_headers(u)

        current_behaviors_dependent_questionnaire = {'dependent_verbal_ability': 'verbal, AACsystem',
                                                     'has_academic_difficulties': False, 'participant_id': p.id}
        rv = self.app.post('api/flow/dependent_intake/current_behaviors_dependent_questionnaire',
                           data=json.dumps(current_behaviors_dependent_questionnaire),
                           content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['dependent_verbal_ability'], 'verbal, AACsystem')
        self.assertEqual(response['has_academic_difficulties'], False)
        self.assertIsNotNone(response['id'])

    def test_current_behaviors_self_questionnaire_basics(self):
        self.construct_current_behaviors_self_questionnaire()
        cbsq = db.session.query(CurrentBehaviorsSelfQuestionnaire).first()
        self.assertIsNotNone(cbsq)
        cbsq_id = cbsq.id
        rv = self.app.get('/api/q/current_behaviors_self_questionnaire/%i' % cbsq_id,
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], cbsq_id)
        self.assertEqual(response["has_academic_difficulties"], cbsq.has_academic_difficulties)

    def test_modify_current_behaviors_self_questionnaire_basics(self):
        self.construct_current_behaviors_self_questionnaire()
        cbsq = db.session.query(CurrentBehaviorsSelfQuestionnaire).first()
        self.assertIsNotNone(cbsq)
        cbsq_id = cbsq.id
        rv = self.app.get('/api/q/current_behaviors_self_questionnaire/%i' % cbsq_id, content_type="application/json",
                           headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['self_verbal_ability'] = ['nonVerbal']
        response['academic_difficulty_areas'] = ['math']
        response['has_academic_difficulties'] = False
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/current_behaviors_self_questionnaire/%i' % cbsq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True, headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/current_behaviors_self_questionnaire/%i' % cbsq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['self_verbal_ability'], ['nonVerbal'])
        self.assertEqual(response['academic_difficulty_areas'], ['math'])
        self.assertEqual(response['has_academic_difficulties'], False)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_current_behaviors_self_questionnaire(self):
        cbsq = self.construct_current_behaviors_self_questionnaire()
        cbsq_id = cbsq.id
        rv = self.app.get('api/q/current_behaviors_self_questionnaire/%i' % cbsq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/current_behaviors_self_questionnaire/%i' % cbsq_id, content_type="application/json"
                             , headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/current_behaviors_self_questionnaire/%i' % cbsq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_current_behaviors_self_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        current_behaviors_self_questionnaire = {'self_verbal_ability': ['verbal', 'AACsystem'],
                                                'has_academic_difficulties': False, 'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/current_behaviors_self_questionnaire',
                           data=json.dumps(current_behaviors_self_questionnaire),
                           content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['self_verbal_ability'], ['verbal', 'AACsystem'])
        self.assertEqual(response['has_academic_difficulties'], False)
        self.assertIsNotNone(response['id'])

    def test_demographics_questionnaire_basics(self):
        self.construct_demographics_questionnaire()
        dq = db.session.query(DemographicsQuestionnaire).first()
        self.assertIsNotNone(dq)
        dq_id = dq.id
        rv = self.app.get('/api/q/demographics_questionnaire/%i' % dq_id,
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], dq_id)
        self.assertEqual(response["birth_sex"], dq.birth_sex)
        self.assertEqual(response["gender_identity"], dq.gender_identity)

    def test_modify_demographics_questionnaire_basics(self):
        self.construct_demographics_questionnaire()
        dq = db.session.query(DemographicsQuestionnaire).first()
        self.assertIsNotNone(dq)
        dq_id = dq.id
        rv = self.app.get('/api/q/demographics_questionnaire/%i' % dq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['gender_identity'] = 'genderOther'
        response['race_ethnicity'] = ['raceOther']
        u2 = self.construct_user(email="rainbows@rainy.com")
        response['user_id'] = u2.id
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/demographics_questionnaire/%i' % dq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True, headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/demographics_questionnaire/%i' % dq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['gender_identity'], 'genderOther')
        self.assertEqual(response['race_ethnicity'], ['raceOther'])
        self.assertEqual(response['user_id'], u2.id)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_demographics_questionnaire(self):
        dq = self.construct_demographics_questionnaire()
        dq_id = dq.id
        rv = self.app.get('api/q/demographics_questionnaire/%i' % dq_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/demographics_questionnaire/%i' % dq_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/demographics_questionnaire/%i' % dq_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_demographics_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        demographics_questionnaire = {'birth_sex': "female", 'gender_identity': "genderOther", 'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/demographics_questionnaire',
                           data=json.dumps(demographics_questionnaire),
                           content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['birth_sex'], 'female')
        self.assertEqual(response['gender_identity'], 'genderOther')
        self.assertIsNotNone(response['id'])

    def test_developmental_questionnaire_basics(self):
        self.construct_developmental_questionnaire()
        dq = db.session.query(DevelopmentalQuestionnaire).first()
        self.assertIsNotNone(dq)
        dq_id = dq.id
        rv = self.app.get('/api/q/developmental_questionnaire/%i' % dq_id,
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], dq_id)
        self.assertEqual(response["had_birth_complications"], dq.had_birth_complications)
        self.assertEqual(response["when_motor_milestones"], dq.when_motor_milestones)

    def test_modify_developmental_questionnaire_basics(self):
        self.construct_developmental_questionnaire()
        dq = db.session.query(DevelopmentalQuestionnaire).first()
        self.assertIsNotNone(dq)
        dq_id = dq.id
        rv = self.app.get('/api/q/developmental_questionnaire/%i' % dq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['when_motor_milestones'] = 'notYet'
        response['when_language_milestones'] = 'notYet'
        response['when_toileting_milestones'] = 'early'
        u2 = self.construct_user(email="rainbows@rainy.com")
        response['user_id'] = u2.id
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/developmental_questionnaire/%i' % dq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True, headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/developmental_questionnaire/%i' % dq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['when_motor_milestones'], 'notYet')
        self.assertEqual(response['when_language_milestones'], 'notYet')
        self.assertEqual(response['when_toileting_milestones'], 'early')
        self.assertEqual(response['user_id'], u2.id)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_developmental_questionnaire(self):
        dq = self.construct_developmental_questionnaire()
        dq_id = dq.id
        rv = self.app.get('api/q/developmental_questionnaire/%i' % dq_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/developmental_questionnaire/%i' % dq_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/developmental_questionnaire/%i' % dq_id, content_type="application/json", headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_developmental_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.dependent)
        headers = self.logged_in_headers(u)

        developmental_questionnaire = {'had_birth_complications': True, 'birth_complications_description': 'C-Section',
                                       'participant_id': p.id}
        rv = self.app.post('api/flow/dependent_intake/developmental_questionnaire',
                           data=json.dumps(developmental_questionnaire),
                           content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['had_birth_complications'], True)
        self.assertEqual(response['birth_complications_description'], 'C-Section')
        self.assertIsNotNone(response['id'])

    def test_education_dependent_questionnaire_basics(self):
        self.construct_education_dependent_questionnaire()
        eq = db.session.query(EducationDependentQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.app.get('/api/q/education_dependent_questionnaire/%i' % eq_id,
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], eq_id)
        self.assertEqual(response["school_name"], eq.school_name)
        self.assertEqual(response["school_type"], eq.school_type)

    def test_modify_education_dependent_questionnaire_basics(self):
        self.construct_education_dependent_questionnaire()
        eq = db.session.query(EducationDependentQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.app.get('/api/q/education_dependent_questionnaire/%i' % eq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['school_name'] = 'Sesame School'
        response['school_type'] = 'public'
        response['dependent_placement'] = 'vocational'
        u2 = self.construct_user(email="rainbows@rainy.com")
        response['user_id'] = u2.id
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/education_dependent_questionnaire/%i' % eq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True, headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/education_dependent_questionnaire/%i' % eq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['school_name'], 'Sesame School')
        self.assertEqual(response['school_type'], 'public')
        self.assertEqual(response['dependent_placement'], 'vocational')
        self.assertEqual(response['user_id'], u2.id)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_education_dependent_questionnaire(self):
        eq = self.construct_education_dependent_questionnaire()
        eq_id = eq.id
        rv = self.app.get('api/q/education_dependent_questionnaire/%i' % eq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/education_dependent_questionnaire/%i' % eq_id, content_type="application/json"
                             , headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/education_dependent_questionnaire/%i' % eq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_education_dependent_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        education_dependent_questionnaire = {'attends_school': True, 'school_name': 'Attreyu Academy',
                                             'participant_id': p.id}
        rv = self.app.post('api/flow/dependent_intake/education_dependent_questionnaire',
                           data=json.dumps(education_dependent_questionnaire),
                           content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['attends_school'], True)
        self.assertEqual(response['school_name'], 'Attreyu Academy')
        self.assertIsNotNone(response['id'])

    def test_education_self_questionnaire_basics(self):
        self.construct_education_self_questionnaire()
        eq = db.session.query(EducationSelfQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.app.get('/api/q/education_self_questionnaire/%i' % eq_id,
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], eq_id)
        self.assertEqual(response["school_name"], eq.school_name)
        self.assertEqual(response["school_type"], eq.school_type)

    def test_modify_education_self_questionnaire_basics(self):
        self.construct_education_self_questionnaire()
        eq = db.session.query(EducationSelfQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.app.get('/api/q/education_self_questionnaire/%i' % eq_id, content_type="application/json"
                          , headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['school_name'] = 'Sesame School'
        response['school_type'] = 'public'
        response['self_placement'] = 'vocational'
        u2 = self.construct_user(email="rainbows@rainy.com")
        response['user_id'] = u2.id
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/education_self_questionnaire/%i' % eq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/education_self_questionnaire/%i' % eq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['school_name'], 'Sesame School')
        self.assertEqual(response['school_type'], 'public')
        self.assertEqual(response['self_placement'], 'vocational')
        self.assertEqual(response['user_id'], u2.id)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_education_self_questionnaire(self):
        eq = self.construct_education_self_questionnaire()
        eq_id = eq.id
        rv = self.app.get('api/q/education_self_questionnaire/%i' % eq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/education_self_questionnaire/%i' % eq_id, content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/education_self_questionnaire/%i' % eq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_education_self_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        education_self_questionnaire = {'attends_school': True, 'school_name': 'Attreyu Academy',
                                        'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/education_self_questionnaire',
                           data=json.dumps(education_self_questionnaire),
                           content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['attends_school'], True)
        self.assertEqual(response['school_name'], 'Attreyu Academy')
        self.assertIsNotNone(response['id'])

    def test_employment_questionnaire_basics(self):
        self.construct_employment_questionnaire()
        eq = db.session.query(EmploymentQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.app.get('/api/q/employment_questionnaire/%i' % eq_id,
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], eq_id)
        self.assertEqual(response["is_currently_employed"], eq.is_currently_employed)
        self.assertEqual(response["employment_capacity"], eq.employment_capacity)
        self.assertEqual(response["has_employment_support"], eq.has_employment_support)

    def test_modify_employment_questionnaire_basics(self):
        self.construct_employment_questionnaire()
        eq = db.session.query(EmploymentQuestionnaire).first()
        self.assertIsNotNone(eq)
        eq_id = eq.id
        rv = self.app.get('/api/q/employment_questionnaire/%i' % eq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['is_currently_employed'] = False
        response['employment_capacity'] = None
        response['has_employment_support'] = 'yes'
        u2 = self.construct_user(email="rainbows@rainy.com")
        response['user_id'] = u2.id
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/employment_questionnaire/%i' % eq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/employment_questionnaire/%i' % eq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['is_currently_employed'], False)
        self.assertEqual(response['employment_capacity'], None)
        self.assertEqual(response['has_employment_support'], 'yes')
        self.assertEqual(response['user_id'], u2.id)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_employment_questionnaire(self):
        eq = self.construct_employment_questionnaire()
        eq_id = eq.id
        rv = self.app.get('api/q/employment_questionnaire/%i' % eq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/employment_questionnaire/%i' % eq_id, content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/employment_questionnaire/%i' % eq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_employment_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        employment_questionnaire = {'is_currently_employed': True, 'employment_capacity': 'partTime',
                                    'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/employment_questionnaire', data=json.dumps(employment_questionnaire),
                           content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['is_currently_employed'], True)
        self.assertEqual(response['employment_capacity'], 'partTime')
        self.assertIsNotNone(response['id'])

    def test_evaluation_history_dependent_questionnaire_basics(self):
        self.construct_evaluation_history_dependent_questionnaire()
        ehq = db.session.query(EvaluationHistoryDependentQuestionnaire).first()
        self.assertIsNotNone(ehq)
        ehq_id = ehq.id
        rv = self.app.get('/api/q/evaluation_history_dependent_questionnaire/%i' % ehq_id,
                          follow_redirects=True,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], ehq_id)
        self.assertEqual(response["self_identifies_autistic"], ehq.self_identifies_autistic)
        self.assertEqual(response["years_old_at_first_diagnosis"], ehq.years_old_at_first_diagnosis)

    def test_modify_evaluation_history_dependent_questionnaire_basics(self):
        self.construct_evaluation_history_dependent_questionnaire()
        ehq = db.session.query(EvaluationHistoryDependentQuestionnaire).first()
        self.assertIsNotNone(ehq)
        ehq_id = ehq.id
        rv = self.app.get('/api/q/evaluation_history_dependent_questionnaire/%i' % ehq_id,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['self_identifies_autistic'] = False
        response['years_old_at_first_diagnosis'] = 12
        response['who_diagnosed'] = 'healthTeam'
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/evaluation_history_dependent_questionnaire/%i' % ehq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/evaluation_history_dependent_questionnaire/%i' % ehq_id,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['self_identifies_autistic'], False)
        self.assertEqual(response['years_old_at_first_diagnosis'], 12)
        self.assertEqual(response['who_diagnosed'], 'healthTeam')
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_evaluation_history_dependent_questionnaire(self):
        ehq = self.construct_evaluation_history_dependent_questionnaire()
        ehq_id = ehq.id
        rv = self.app.get('api/q/evaluation_history_dependent_questionnaire/%i' % ehq_id,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/evaluation_history_dependent_questionnaire/%i' % ehq_id,
                             content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/evaluation_history_dependent_questionnaire/%i' % ehq_id,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_evaluation_history_dependent_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_guardian)
        headers = self.logged_in_headers(u)

        evaluation_history_dependent_questionnaire = {'self_identifies_autistic': True,
                                                      'years_old_at_first_diagnosis': 5,
                                                      'participant_id': p.id}
        rv = self.app.post('api/flow/dependent_intake/evaluation_history_dependent_questionnaire',
                           data=json.dumps(evaluation_history_dependent_questionnaire), content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['self_identifies_autistic'], True)
        self.assertEqual(response['years_old_at_first_diagnosis'], 5)
        self.assertIsNotNone(response['id'])

    def test_evaluation_history_self_questionnaire_basics(self):
        self.construct_evaluation_history_self_questionnaire()
        ehq = db.session.query(EvaluationHistorySelfQuestionnaire).first()
        self.assertIsNotNone(ehq)
        ehq_id = ehq.id
        rv = self.app.get('/api/q/evaluation_history_self_questionnaire/%i' % ehq_id,
                          follow_redirects=True,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], ehq_id)
        self.assertEqual(response["self_identifies_autistic"], ehq.self_identifies_autistic)
        self.assertEqual(response["years_old_at_first_diagnosis"], ehq.years_old_at_first_diagnosis)

    def test_modify_evaluation_history_self_questionnaire_basics(self):
        self.construct_evaluation_history_self_questionnaire()
        ehq = db.session.query(EvaluationHistorySelfQuestionnaire).first()
        self.assertIsNotNone(ehq)
        ehq_id = ehq.id
        rv = self.app.get('/api/q/evaluation_history_self_questionnaire/%i' % ehq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['self_identifies_autistic'] = False
        response['years_old_at_first_diagnosis'] = 12
        response['who_diagnosed'] = 'healthTeam'
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/evaluation_history_self_questionnaire/%i' % ehq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/evaluation_history_self_questionnaire/%i' % ehq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['self_identifies_autistic'], False)
        self.assertEqual(response['years_old_at_first_diagnosis'], 12)
        self.assertEqual(response['who_diagnosed'], 'healthTeam')
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_evaluation_history_self_questionnaire(self):
        ehq = self.construct_evaluation_history_self_questionnaire()
        ehq_id = ehq.id
        rv = self.app.get('api/q/evaluation_history_self_questionnaire/%i' % ehq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/evaluation_history_self_questionnaire/%i' % ehq_id, content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/evaluation_history_self_questionnaire/%i' % ehq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_evaluation_history_self_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_guardian)
        headers = self.logged_in_headers(u)

        evaluation_history_self_questionnaire = {'self_identifies_autistic': True, 'years_old_at_first_diagnosis': 5,
                                                 'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/evaluation_history_self_questionnaire',
                           data=json.dumps(evaluation_history_self_questionnaire), content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['self_identifies_autistic'], True)
        self.assertEqual(response['years_old_at_first_diagnosis'], 5)
        self.assertIsNotNone(response['id'])

    def test_home_dependent_questionnaire_basics(self):
        self.construct_home_dependent_questionnaire()
        hq = db.session.query(HomeDependentQuestionnaire).first()
        self.assertIsNotNone(hq)
        hq_id = hq.id
        rv = self.app.get('/api/q/home_dependent_questionnaire/%i' % hq_id,
                          follow_redirects=True,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], hq_id)
        self.assertEqual(response["participant_id"], hq.participant_id)
        self.assertEqual(response["user_id"], hq.user_id)
        self.assertEqual(response["dependent_living_situation"], hq.dependent_living_situation)
        self.assertEqual(response["struggle_to_afford"], hq.struggle_to_afford)
        self.assertEqual(len(response["housemates"]), len(hq.housemates))

    def test_modify_home_dependent_questionnaire_basics(self):
        self.construct_home_dependent_questionnaire()
        hq = db.session.query(HomeDependentQuestionnaire).first()
        self.assertIsNotNone(hq)
        hq_id = hq.id
        rv = self.app.get('/api/q/home_dependent_questionnaire/%i' % hq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['participant_id'] = self.construct_participant(user=self.construct_user(),
                                                                relationship=Relationship.dependent).id
        response['dependent_living_situation'] = ['caregiver']
        response['struggle_to_afford'] = True
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/home_dependent_questionnaire/%i' % hq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        self.construct_housemate(name='Debbie Danger', home_dependent_questionnaire=hq)
        rv = self.app.get('/api/q/home_dependent_questionnaire/%i' % hq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['dependent_living_situation'], ['caregiver'])
        self.assertEqual(response['struggle_to_afford'], True)
        self.assertEqual(len(response['housemates']), 2)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_home_dependent_questionnaire(self):
        hq = self.construct_home_dependent_questionnaire()
        hq_id = hq.id
        rv = self.app.get('api/q/home_dependent_questionnaire/%i' % hq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/home_dependent_questionnaire/%i' % hq_id, content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/home_dependent_questionnaire/%i' % hq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_home_dependent_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        home_dependent_questionnaire = {'dependent_living_situation': ['family'], 'struggle_to_afford': False,
                                        'participant_id': p.id}
        rv = self.app.post('api/flow/dependent_intake/home_dependent_questionnaire',
                           data=json.dumps(home_dependent_questionnaire), content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['participant_id'], p.id)
        self.assertEqual(response['dependent_living_situation'], ['family'])
        self.assertEqual(response['struggle_to_afford'], False)
        self.assertIsNotNone(response['id'])

    def test_home_self_questionnaire_basics(self):
        self.construct_home_self_questionnaire()
        hq = db.session.query(HomeSelfQuestionnaire).first()
        self.assertIsNotNone(hq)
        hq_id = hq.id
        rv = self.app.get('/api/q/home_self_questionnaire/%i' % hq_id,
                          follow_redirects=True,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], hq_id)
        self.assertEqual(response["participant_id"], hq.participant_id)
        self.assertEqual(response["user_id"], hq.user_id)
        self.assertEqual(response["self_living_situation"], hq.self_living_situation)
        self.assertEqual(response["struggle_to_afford"], hq.struggle_to_afford)
        self.assertEqual(len(response["housemates"]), len(hq.housemates))

    def test_modify_home_self_questionnaire_basics(self):
        self.construct_home_self_questionnaire()
        hq = db.session.query(HomeSelfQuestionnaire).first()
        self.assertIsNotNone(hq)
        hq_id = hq.id
        rv = self.app.get('/api/q/home_self_questionnaire/%i' % hq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['participant_id'] = self.construct_participant(user=self.construct_user(),
                                                                relationship=Relationship.self_participant).id
        response['self_living_situation'] = ['caregiver']
        response['struggle_to_afford'] = True
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/home_self_questionnaire/%i' % hq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        self.construct_housemate(name='Debbie Danger', home_self_questionnaire=hq)
        rv = self.app.get('/api/q/home_self_questionnaire/%i' % hq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['self_living_situation'], ['caregiver'])
        self.assertEqual(response['struggle_to_afford'], True)
        self.assertEqual(len(response['housemates']), 2)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_home_self_questionnaire(self):
        hq = self.construct_home_self_questionnaire()
        hq_id = hq.id
        rv = self.app.get('api/q/home_self_questionnaire/%i' % hq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/home_self_questionnaire/%i' % hq_id, content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/home_self_questionnaire/%i' % hq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_home_self_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        home_self_questionnaire = {'self_living_situation': ['family'], 'struggle_to_afford': False,
                                   'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/home_self_questionnaire',
                           data=json.dumps(home_self_questionnaire), content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['participant_id'], p.id)
        self.assertEqual(response['self_living_situation'], ['family'])
        self.assertEqual(response['struggle_to_afford'], False)
        self.assertIsNotNone(response['id'])

    def test_identification_questionnaire_basics(self):
        self.construct_identification_questionnaire()
        iq = db.session.query(IdentificationQuestionnaire).first()
        self.assertIsNotNone(iq)
        iq_id = iq.id
        rv = self.app.get('/api/q/identification_questionnaire/%i' % iq_id,
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], iq_id)
        self.assertEqual(response["first_name"], iq.first_name)
        self.assertEqual(response["nickname"], iq.nickname)
        self.assertEqual(response["birth_state"], iq.birth_state)

    def test_modify_identification_questionnaire_basics(self):
        self.construct_identification_questionnaire()
        iq = db.session.query(IdentificationQuestionnaire).first()
        self.assertIsNotNone(iq)
        iq_id = iq.id
        rv = self.app.get('/api/q/identification_questionnaire/%i' % iq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['first_name'] = 'Helga'
        response['birth_city'] = 'Staunton'
        response['is_first_name_preferred'] = True
        u2 = self.construct_user(email="rainbows@rainy.com")
        response['user_id'] = u2.id
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/identification_questionnaire/%i' % iq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        rv = self.app.get('/api/q/identification_questionnaire/%i' % iq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['first_name'], 'Helga')
        self.assertEqual(response['birth_city'], 'Staunton')
        self.assertEqual(response['is_first_name_preferred'], True)
        self.assertEqual(response['user_id'], u2.id)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_identification_questionnaire(self):
        iq = self.construct_identification_questionnaire()
        iq_id = iq.id
        rv = self.app.get('api/q/identification_questionnaire/%i' % iq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/identification_questionnaire/%i' % iq_id, content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/identification_questionnaire/%i' % iq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_identification_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        identification_questionnaire = {'first_name': 'Eloise', 'middle_name': 'Elora', 'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/identification_questionnaire',
                           data=json.dumps(identification_questionnaire),
                           content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['first_name'], 'Eloise')
        self.assertEqual(response['middle_name'], 'Elora')
        self.assertIsNotNone(response['id'])

    def test_supports_questionnaire_basics(self):
        self.construct_supports_questionnaire()
        sq = db.session.query(SupportsQuestionnaire).first()
        self.assertIsNotNone(sq)
        sq_id = sq.id
        rv = self.app.get('/api/q/supports_questionnaire/%i' % sq_id,
                          follow_redirects=True,
                          content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response["id"], sq_id)
        self.assertEqual(response["participant_id"], sq.participant_id)
        self.assertEqual(response["user_id"], sq.user_id)
        self.assertEqual(len(response["assistive_devices"]), len(sq.assistive_devices))
        self.assertEqual(len(response["medications"]), len(sq.medications))
        self.assertEqual(len(response["therapies"]), len(sq.therapies))

    def test_modify_supports_questionnaire_basics(self):
        self.construct_supports_questionnaire()
        sq = db.session.query(SupportsQuestionnaire).first()
        self.assertIsNotNone(sq)
        sq_id = sq.id
        rv = self.app.get('/api/q/supports_questionnaire/%i' % sq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        response = json.loads(rv.get_data(as_text=True))
        response['participant_id'] = self.construct_participant(user=self.construct_user(),
                                                                relationship=Relationship.self_participant).id
        orig_date = response['last_updated']
        rv = self.app.put('/api/q/supports_questionnaire/%i' % sq_id, data=json.dumps(response),
                          content_type="application/json",
                          follow_redirects=True,
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        self.construct_medication(name='Iocane Powder', supports_questionnaire=sq)
        self.construct_therapy(type='socialSkills', supports_questionnaire=sq)
        self.construct_alternative_augmentative(type='highTechAAC', supports_questionnaire=sq)
        self.construct_assistive_device(type_group='hearing', type='hearingAid', notes='Your ears you keep and I\'ll tell you why.', supports_questionnaire=sq)
        rv = self.app.get('/api/q/supports_questionnaire/%i' % sq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(len(response['medications']), 2)
        self.assertEqual(len(response['therapies']), 2)
        self.assertEqual(len(response['alternative_augmentative']), 2)
        self.assertEqual(len(response['assistive_devices']), 2)
        self.assertNotEqual(orig_date, response['last_updated'])

    def test_delete_supports_questionnaire(self):
        sq = self.construct_supports_questionnaire()
        sq_id = sq.id
        rv = self.app.get('api/q/supports_questionnaire/%i' % sq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.delete('api/q/supports_questionnaire/%i' % sq_id, content_type="application/json",
                             headers=self.logged_in_headers())
        self.assertSuccess(rv)

        rv = self.app.get('api/q/supports_questionnaire/%i' % sq_id, content_type="application/json",
                          headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code)

    def test_create_supports_questionnaire(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        supports_questionnaire = {'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/supports_questionnaire',
                           data=json.dumps(supports_questionnaire), content_type="application/json",
                           follow_redirects=True,
                           headers=headers)
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(response['participant_id'], p.id)
        self.assertIsNotNone(response['id'])

    def test_questionnare_post_fails_if_flow_does_not_exist(self):
        evaluation_history_self_questionnaire = {'self_identifies_autistic': True, 'years_old_at_first_diagnosis': 5}
        rv = self.app.post('api/flow/noSuchFlow/evaluation_history_self_questionnaire',
                           data=json.dumps(evaluation_history_self_questionnaire), content_type="application/json",
                           follow_redirects=True,
                           headers=self.logged_in_headers())
        self.assertEqual(404, rv.status_code,
                         "This endpoint should require that the flow exists and that the question is in the flow.")
        response = json.loads(rv.get_data(as_text=True))
        self.assertIsNotNone(response)
        self.assertEqual("Unknown path.", response["message"],
                         "There should be a clear error message explaining what went wrong.")

    def test_questionnare_post_fails_if_question_not_in_flow(self):
        evaluation_history_self_questionnaire = {'self_identifies_autistic': True, 'years_old_at_first_diagnosis': 5}
        rv = self.app.post('api/flow/self_intake/guardian_demographics_questionnaire',
                           data=json.dumps(evaluation_history_self_questionnaire), content_type="application/json",
                           follow_redirects=True,
                           headers=self.logged_in_headers())
        self.assertEqual(400, rv.status_code,
                         "This endpoint should require that the flow exists and that the question is in the flow.")
        response = json.loads(rv.get_data(as_text=True))
        self.assertIsNotNone(response)
        self.assertEqual("not_in_the_flow", response["code"],
                         "There should be a clear error message explaining what went wrong.")

    def test_questionnare_post_fails_if_not_logged_in(self):
        cq = {'first_name': "Darah", 'marketing_channel': "Subway sign"}
        rv = self.app.post('api/flow/self_intake/contact_questionnaire', data=json.dumps(cq),
                           content_type="application/json",
                           follow_redirects=True)
        self.assertEqual(401, rv.status_code)
        pass

    def test_questionnaire_post_fails_if_user_not_connected_to_participant(self):
        cq = {'first_name': "Darah", 'marketing_channel': "Subway sign"}
        rv = self.app.post('api/flow/self_intake/contact_questionnaire', data=json.dumps(cq),
                           content_type="application/json",
                           follow_redirects=True, headers=self.logged_in_headers())
        self.assertEqual(400, rv.status_code,
                         "This endpoint should require a participant id that is associated with current user.")
        response = json.loads(rv.get_data(as_text=True))
        self.assertIsNotNone(response)
        self.assertEqual("Unable to save the provided object.", response["message"],
                         "There should be a clear error message explaining what went wrong.")

    def test_questionnionare_post_creates_log_record(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)

        cq = {'first_name': "Darah", 'marketing_channel': "Subway sign", 'participant_id': p.id}
        rv = self.app.post('api/flow/self_intake/contact_questionnaire', data=json.dumps(cq),
                           content_type="application/json",
                           follow_redirects=True, headers=headers)
        self.assertSuccess(rv)
        log = db.session.query(StepLog).all()
        self.assertIsNotNone(log)
        self.assertTrue(len(log) > 0)

    def test_flow_endpoint(self):
        # It should be possible to get a list of available flows
        rv = self.app.get('api/flow', content_type="application/json", headers=self.logged_in_headers())
        self.assertEqual(200, rv.status_code)
        response = json.loads(rv.get_data(as_text=True))
        self.assertIsNotNone(response)
        self.assertTrue(len(response) > 0)

    def test_intake_flows_endpoint(self):
        # Are the basics correct about the existing intake flows?
        rv = self.app.get('api/flow', content_type="application/json", headers=self.logged_in_headers())
        self.assertEqual(200, rv.status_code)
        response = json.loads(rv.get_data(as_text=True))
        for i in response:
            if i['name'] == 'self_intake':
                self.assertEqual(len(i['steps']), 10)
                self.assertEqual(i['steps'][8]['name'], 'employment_questionnaire')
                self.assertEqual(i['steps'][8]['label'], 'Employment')
            if i['name'] == 'dependent_intake':
                self.assertEqual(len(i['steps']), 9)
                self.assertEqual(i['steps'][5]['name'], 'developmental_questionnaire')
                self.assertEqual(i['steps'][5]['label'], 'Birth and Developmental History')
            if i['name'] == 'guardian_intake':
                self.assertEqual(len(i['steps']), 3)
                self.assertEqual(i['steps'][1]['name'], 'contact_questionnaire')
                self.assertEqual(i['steps'][1]['label'], 'Contact Information')

    def test_self_intake_flow_with_user(self):
        u = self.construct_user()
        p = self.construct_participant(user=u, relationship=Relationship.self_participant)
        headers = self.logged_in_headers(u)
        rv = self.app.get('api/flow/self_intake/%i' % p.id, content_type="application/json", headers=headers)
        self.assertEqual(200, rv.status_code)
        response = json.loads(rv.get_data(as_text=True))
        self.assertIsNotNone(response)
        self.assertEqual('self_intake', response['name'])
        self.assertIsNotNone(response['steps'])
        self.assertTrue(len(response['steps']) > 0)
        self.assertEqual('identification_questionnaire', response['steps'][0]['name'])
        self.assertEqual(QuestionService.TYPE_IDENTIFYING, response['steps'][0]['type'])
        self.assertEqual(Step.STATUS_INCOMPLETE, response['steps'][0]['status'])

        cq = {
            'first_name': "Darah",
            'middle_name': "Soo",
            'last_name': "Ubway",
            'is_first_name_preferred': True,
            'birthdate': '02/02/2002',
            'birth_city': 'Staunton',
            'birth_state': 'VA',
            'is_english_primary': True,
            'participant_id': p.id
        }
        rv = self.app.post('api/flow/self_intake/identification_questionnaire', data=json.dumps(cq),
                           content_type="application/json",
                           follow_redirects=True, headers=headers)

        rv = self.app.get('api/flow/self_intake/%i' % p.id, content_type="application/json", headers=headers)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual('identification_questionnaire', response['steps'][0]['name'])
        self.assertEqual(Step.STATUS_COMPLETE, response['steps'][0]['status'])
        self.assertIsNotNone(response['steps'][0]['date_completed'])

    def test_questionnaire_meta_is_relation_specific(self):
        self.construct_identification_questionnaire()
        rv = self.app.get('/api/flow/self_intake/identification_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        intro = self.get_field_from_response(response, "intro")
        self.assertIsNotNone(intro["template_options"]["description"])
        self.assertEqual(intro["template_options"]["description"],
                         "Please answer the following questions about yourself (* indicates required response):")

        birth_city = self.get_field_from_response(response, "birth_city")
        self.assertIsNotNone(birth_city)
        self.assertIsNotNone(birth_city["template_options"])
        self.assertIsNotNone(birth_city["template_options"]["label"])
        self.assertEqual(birth_city["template_options"]["label"],
                         "Your city/municipality of birth")

        rv = self.app.get('/api/flow/dependent_intake/identification_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        birth_city = self.get_field_from_response(response, "birth_city")
        self.assertIsNotNone(birth_city["template_options"]["label"])
        self.assertEqual(birth_city["template_options"]["label"],
                         "Your child's city/municipality of birth")

    def test_questionnaire_meta_has_relation_required_fields(self):
        self.construct_identification_questionnaire()
        rv = self.app.get('/api/flow/dependent_intake/identification_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        relationship = self.get_field_from_response(response, "relationship")
        self.assertIsNotNone(relationship)

        # Convert Participant to a dependant
        rv = self.app.get('/api/flow/self_intake/identification_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        relationship = self.get_field_from_response(response, "relationship")
        self.assertIsNone(relationship)

    def test_meta_contains_table_details(self):
        self.construct_identification_questionnaire()
        rv = self.app.get('/api/flow/dependent_intake/identification_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual("identifying", response["table"]["question_type"])
        self.assertEqual("Identification", response["table"]["label"])

    def test_meta_field_groups_contain_their_fields(self):
        self.construct_home_self_questionnaire()
        rv = self.app.get('/api/flow/self_intake/home_self_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self_living = self.get_field_from_response(response, "self_living")
        self.assertEqual("self_living_situation", self_living["fieldGroup"][0]["name"])

    def test_support_meta_contain_their_fields(self):
        self.construct_supports_questionnaire()
        rv = self.app.get('/api/flow/self_intake/supports_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        assistive_devices = self.get_field_from_response(response, "assistive_devices")
        self.assertIsNotNone(assistive_devices["fieldArray"]["fieldGroup"][0])
        self.assertEqual("type", assistive_devices["fieldArray"]["fieldGroup"][0]["name"])

    def test_evaluation_history_dependent_meta_contain_their_fields(self):
        self.construct_evaluation_history_dependent_questionnaire()
        rv = self.app.get('/api/flow/dependent_intake/evaluation_history_dependent_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(len(response["fields"]), 10)

    def test_evaluation_history_self_meta_contain_their_fields(self):
        self.construct_evaluation_history_self_questionnaire()
        rv = self.app.get('/api/flow/self_intake/evaluation_history_self_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(len(response["fields"]), 10)

    def test_education_dependent_meta_contain_their_fields(self):
        self.construct_education_dependent_questionnaire()
        rv = self.app.get('/api/flow/dependent_intake/education_dependent_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(len(response["fields"]), 5)

    def test_education_self_meta_contain_their_fields(self):
        self.construct_education_self_questionnaire()
        rv = self.app.get('/api/flow/self_intake/education_self_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(len(response["fields"]), 5)

    def test_meta_fields_are_ordered(self):
        self.construct_education_self_questionnaire()
        rv = self.app.get('/api/flow/self_intake/education_self_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(1, response["fields"][0]["display_order"])
        self.assertEqual(2, response["fields"][1]["display_order"])
        self.assertEqual(3, response["fields"][2]["display_order"])

        self.assertEqual(6.1, response['fields'][4]["fieldGroup"][0]["display_order"])
        self.assertEqual(6.2, response['fields'][4]["fieldGroup"][1]["display_order"])

    def test_questionnaire_names_list_basics(self):
        rv = self.app.get('/api/q',
                          follow_redirects=True,
                          content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual("assistive_device", response[1])
        self.assertEqual("education_dependent_questionnaire", response[8])
        self.assertEqual("home_self_questionnaire", response[14])
        self.assertEqual("therapy", response[20])
        self.assertEqual(21, len(response))

    def test_questionnaire_list_meta_basics(self):
        self.construct_education_self_questionnaire()
        rv = self.app.get('/api/q/education_self_questionnaire/meta',
                          follow_redirects=True,
                          content_type="application/json")
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual("id", response["fields"][0]["name"])
        self.assertEqual("user_id", response["fields"][4]["name"])
        self.assertEqual("school_type", response["fields"][7]["name"])
        self.assertEqual("school_services_other", response["fields"][12]["name"])
        self.assertEqual(13, len(response["fields"]))

    def test_questionnaire_list_basics(self):
        self.construct_current_behaviors_dependent_questionnaire()
        rv = self.app.get('/api/q/current_behaviors_dependent_questionnaire',
                          follow_redirects=True,
                          content_type="application/json", headers=self.logged_in_headers())
        self.assertSuccess(rv)
        response = json.loads(rv.get_data(as_text=True))
        self.assertEqual(["math", "writing"], response[0]["academic_difficulty_areas"])
        self.assertEqual("fluent", response[0]["dependent_verbal_ability"])
        self.assertEqual(1, response[0]["id"])

    def test_non_admin_cannot_view_questionnaire_list(self):
        user = self.construct_user(email='regularUser@user.com')
        admin = self.construct_admin_user(email='adminUser@user.com')
        self.construct_contact_questionnaire()
        rv = self.app.get('/api/q/contact_questionnaire', content_type="application/json")
        self.assertEqual(401, rv.status_code)
        rv = self.app.get('/api/q/contact_questionnaire', content_type="application/json", headers=self.logged_in_headers(user=user))
        self.assertEqual(403, rv.status_code)
        rv = self.app.get('/api/q/contact_questionnaire', content_type="application/json", headers=self.logged_in_headers(user=admin))
        self.assertSuccess(rv)

    def test_export_single_questionnaire(self):
        self.construct_current_behaviors_dependent_questionnaire()
        rv = self.app.get('/api/q/current_behaviors_dependent_questionnaire/export',
                          follow_redirects=True,
                          content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        wb = openpyxl.load_workbook(io.BytesIO(rv.data))
        ws = wb.get_active_sheet()
        self.assertEqual(ws, wb.active)
        self.assertEqual('id', ws['A1'].value)
        self.assertEqual('last_updated', ws['B1'].value)
        self.assertEqual('participant_id', ws['C1'].value)
        self.assertEqual('user_id', ws['D1'].value)
        self.assertEqual('dependent_verbal_ability', ws['E1'].value)
        self.assertEqual('concerning_behaviors', ws['F1'].value)
        self.assertEqual('hoarding, ', ws['F2'].value)
        self.assertEqual('concerning_behaviors_other', ws['G1'].value)
        self.assertEqual('has_academic_difficulties', ws['H1'].value)
        self.assertEqual('academic_difficulty_areas', ws['I1'].value)
        self.assertEqual('math, writing, ', ws['I2'].value)
        self.assertEqual('academic_difficulty_other', ws['J1'].value)
        self.assertEqual(10, ws.max_column)
        self.assertEqual(2, ws.max_row)

    def test_export_all_questionnaires(self):
        self.construct_all_questionnaires()
        rv = self.app.get('/api/q/all/export', follow_redirects=True,
                          content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          headers=self.logged_in_headers())
        self.assertSuccess(rv)
        wb = openpyxl.load_workbook(io.BytesIO(rv.data))
        ws = wb.get_active_sheet()
        self.assertEqual(ws, wb.active)
        self.assertEqual(2, ws.max_row)
        self.assertEqual(21, len(wb.worksheets))
        self.assertEqual('alternative_augmentative', wb.worksheets[0].title)
        self.assertEqual('assistive_device', wb.worksheets[1].title)
        self.assertEqual('clinical_diagnoses_questionnai', wb.worksheets[2].title)
        self.assertEqual('contact_questionnaire', wb.worksheets[3].title)
        self.assertEqual('current_behaviors_dependent_qu', wb.worksheets[4].title)
        self.assertEqual('current_behaviors_self_questio', wb.worksheets[5].title)
        self.assertEqual('demographics_questionnaire', wb.worksheets[6].title)
        self.assertEqual('developmental_questionnaire', wb.worksheets[7].title)
        self.assertEqual('education_dependent_questionna', wb.worksheets[8].title)
        self.assertEqual('education_self_questionnaire', wb.worksheets[9].title)
        self.assertEqual('employment_questionnaire', wb.worksheets[10].title)
        self.assertEqual('evaluation_history_dependent_q', wb.worksheets[11].title)
        self.assertEqual('evaluation_history_self_questi', wb.worksheets[12].title)
        self.assertEqual('home_dependent_questionnaire', wb.worksheets[13].title)
        self.assertEqual('home_self_questionnaire', wb.worksheets[14].title)
        self.assertEqual('housemate', wb.worksheets[15].title)
        self.assertEqual('identification_questionnaire', wb.worksheets[16].title)
        self.assertEqual('medication', wb.worksheets[17].title)
        self.assertEqual('professional_profile_questionn', wb.worksheets[18].title)
        self.assertEqual('supports_questionnaire', wb.worksheets[19].title)
        self.assertEqual('therapy', wb.worksheets[20].title)
